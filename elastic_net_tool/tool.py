"""ModelingTool — main orchestration class for elastic net insurance GLMs."""

from __future__ import annotations

from collections.abc import Sequence
from dataclasses import replace
from pathlib import Path
from typing import Any, Callable

import matplotlib.pyplot as plt
import numpy as np
import polars as pl
from tabulate import tabulate

from .io_utils import (
    _serialize_custom_transform,
    save_version,
)
from .io_utils import load_version as _load_version_snapshot
from .metrics import (
    _off_balance_by_state,
    compare_metrics,
    compute_metrics,
    compute_midpoint_movement,
    double_lift_score,
    double_lift_table,
    gini_coefficient,
)
from .model import (
    FactorModelVersion,
    ModelVersion,
    _build_preprocessor,
    fit_cv_stability,
    fit_model,
)
from .plots import (
    _resolve_level,
    _sort_labels,
    ae_chart,
    coefficient_plot,
    cv_stability_plot,
    decile_lift_chart,
    double_lift_chart,
    midpoint_movement_histogram,
    residual_chart,
    univariate_plot,
)
from .variable import (
    MISSING_SENTINEL,
    FittedBinnedNumericParams,
    FittedCategoricalParams,
    Preprocessor,
    VariableConfig,
    default_config,
)

# ── Module-level helpers for relativities_table ───────────────────────────────


def _validate_numeric_column(
    data: pl.DataFrame,
    argument_name: str,
    column: str | None,
) -> None:
    """Validate an optional numeric column argument."""
    if column is None:
        return
    if column not in data.columns:
        raise ValueError(f"{argument_name} '{column}' not found in data.")
    if not data[column].dtype.is_numeric():
        raise ValueError(
            f"{argument_name} '{column}' must be numeric, got {data[column].dtype}."
        )


def _validate_fit_request(variables: list[str], version: str) -> None:
    """Validate the public variable-list and version arguments."""
    invalid_variables = (
        not isinstance(variables, list)
        or not variables
        or any(not isinstance(variable, str) or not variable for variable in variables)
    )
    if invalid_variables:
        raise ValueError("variables must be a non-empty list of variable names.")
    if len(set(variables)) != len(variables):
        raise ValueError("variables must not contain duplicates.")
    if not isinstance(version, str) or not version:
        raise ValueError("version must be a non-empty string.")


def _create_variable_config(
    data: pl.DataFrame,
    column: str,
    input_columns: list[str] | None,
    custom_transform: Callable | None,
    options: dict[str, Any],
) -> VariableConfig:
    """Build a variable config from the keyword-oriented public API."""
    if "breakpoints" in options:
        options["bin_edges"] = options.pop("breakpoints")
    if input_columns is not None:
        options["input_cols"] = input_columns
    if custom_transform is not None:
        options["custom_transform"] = custom_transform
    if options:
        return VariableConfig(col=column, **options)
    if column in data.columns:
        return default_config(column, data[column])
    return VariableConfig(col=column)


def _fit_plot_preprocessor(
    col: str,
    data: pl.DataFrame,
    configs: dict[str, VariableConfig],
    weights: np.ndarray | None,
) -> Preprocessor:
    """Fit a preprocessor without rejecting plot-binned numeric missings."""
    preprocessor = _build_preprocessor([col], data, configs)
    config = preprocessor.configs[col]
    if (
        config.impute_strategy is None
        and config.n_bins is None
        and config.bin_edges is None
    ):
        materialized = preprocessor._materialize_raw_columns(
            data,
            [col],
            fit_aliases=True,
            weights=weights,
        )
        raw = materialized[col]
        if raw.dtype.is_numeric():
            arr = (
                raw.cast(pl.Float64, strict=False)
                .fill_null(MISSING_SENTINEL)
                .to_numpy(allow_copy=True)
            )
            missing = ~np.isfinite(arr) | np.isclose(
                arr, MISSING_SENTINEL, rtol=0, atol=1.0
            )
            if missing.any():
                plot_configs = dict(configs)
                plot_configs[col] = replace(
                    config,
                    impute_strategy="constant",
                    impute_value=MISSING_SENTINEL,
                )
                preprocessor = _build_preprocessor([col], data, plot_configs)
    return preprocessor.fit(data, weights=weights)


def _transform_signature(
    config: VariableConfig,
    saved_source: dict[str, Any] | None = None,
) -> dict[str, Any] | None:
    """Return a stable signature for a config's custom transform."""
    if config.custom_transform is None:
        return None
    if saved_source is not None:
        return saved_source
    embedded = getattr(
        config.custom_transform,
        "__elastic_net_tool_transform_source__",
        None,
    )
    return embedded or _serialize_custom_transform(config.custom_transform)


def _configs_match(
    existing: VariableConfig,
    incoming: VariableConfig,
    incoming_source: dict[str, Any] | None,
) -> bool:
    """Compare configs by values and transform source rather than callable identity."""
    existing_plain = replace(existing, custom_transform=None)
    incoming_plain = replace(incoming, custom_transform=None)
    try:
        values_match = existing_plain == incoming_plain
    except ValueError:
        values_match = repr(existing_plain) == repr(incoming_plain)
    return bool(values_match) and _transform_signature(
        existing
    ) == _transform_signature(incoming, incoming_source)


def _validate_load_destination(
    tool: ModelingTool,
    tool_settings: dict[str, Any],
    version_name: str,
) -> None:
    """Reject version, tool-setting, and variable-config conflicts."""
    if version_name in tool.model_versions:
        raise ValueError(f"Model version '{version_name}' already exists.")

    saved_settings = {
        "target_col": tool_settings["target_col"],
        "weight_col": tool_settings["weight_col"],
        "offset_col": tool_settings.get("offset_col"),
        "drop_reference": tool_settings.get("drop_reference", "max_weight"),
    }
    mismatches = [
        setting
        for setting, saved_value in saved_settings.items()
        if getattr(tool, setting) != saved_value
    ]
    if mismatches:
        details = ", ".join(
            f"{setting} (tool={getattr(tool, setting)!r}, saved={saved_settings[setting]!r})"
            for setting in mismatches
        )
        raise ValueError(
            f"Saved model is incompatible with the destination tool: {details}."
        )

    sources = tool_settings.get("custom_transform_sources", {})
    conflicts = [
        col
        for col, incoming in tool_settings["variable_configs"].items()
        if col in tool.variable_configs
        and not _configs_match(tool.variable_configs[col], incoming, sources.get(col))
    ]
    if conflicts:
        raise ValueError(
            "Saved model has incompatible variable configurations for: "
            + ", ".join(conflicts)
            + "."
        )


def _merge_variable_configs(
    tool: ModelingTool,
    tool_settings: dict[str, Any],
) -> dict[str, VariableConfig]:
    """Return the destination configs plus the artifact's compatible configs."""
    configs = dict(tool.variable_configs)
    for col, config in tool_settings["variable_configs"].items():
        configs.setdefault(col, config)
    return configs


def _fit_loaded_snapshot(
    tool: ModelingTool,
    snapshot: dict[str, Any],
    version_name: str,
) -> None:
    """Refit one loaded GLM snapshot on a prepared tool."""
    version = snapshot["version"]
    tool.fit_model(
        variables=version["variables"],
        version=version_name,
        alpha=version["alpha"],
        l1_ratio=version["l1_ratio"],
        use_cv=False,
        family=version["family"],
        tweedie_power=version["tweedie_power"],
        link=version["link"],
        gradient_tol=version.get("gradient_tol"),
        print_summary=True,
    )


def _new_frozen_tool(
    tool_class: type[ModelingTool],
    snapshot: dict[str, Any],
    data: pl.DataFrame | None,
) -> ModelingTool:
    """Create a data-optional tool shell for a frozen artifact."""
    settings = snapshot["tool_settings"]
    tool = tool_class.__new__(tool_class)
    tool.data = data.clone() if data is not None else pl.DataFrame()
    tool.target_col = settings["target_col"]
    tool.weight_col = settings["weight_col"]
    tool.offset_col = settings.get("offset_col")
    tool.link = settings["link"]
    tool.tweedie_power = 1.5
    tool.drop_reference = settings.get("drop_reference", "max_weight")
    tool.cv_column = None
    tool.current_version = None
    tool.base_version = None
    tool.variable_configs = {}
    tool.model_versions = {}
    return tool


def _restore_frozen_model(
    snapshot: dict[str, Any],
    version_name: str,
    data: pl.DataFrame,
    compute_predictions: bool,
) -> ModelVersion | FactorModelVersion:
    """Reconstruct one fitted model and optionally score the tool data."""
    version = snapshot["version"]
    offset_col = snapshot["tool_settings"].get("offset_col")
    offset = None
    if compute_predictions and offset_col is not None and offset_col in data.columns:
        offset = data[offset_col].cast(pl.Float64).to_numpy()

    if snapshot.get("version_type") == "factor":
        model: ModelVersion | FactorModelVersion = FactorModelVersion(
            name=version_name,
            variables=version["variables"],
            factor_table=version["factor_table"],
            preprocessor=version["preprocessor"],
            preprocessor_vars=version["preprocessor_vars"],
            train_predictions=np.array([]),
            offset_col=version["offset_col"],
            fit_info=version.get("fit_info", {}),
        )
    else:
        model = ModelVersion(
            name=version_name,
            variables=version["variables"],
            preprocessor=version["preprocessor"],
            glm=version["glm"],
            feature_names=version["feature_names"],
            coefficients=version["coefficients"],
            alpha=version["alpha"],
            l1_ratio=version["l1_ratio"],
            family=version["family"],
            link=version["link"],
            train_predictions=np.array([]),
            fit_info=version["fit_info"],
            cv_stability=version.get("cv_stability"),
            tweedie_power=version["tweedie_power"],
            gradient_tol=version.get("gradient_tol"),
        )
    if compute_predictions:
        model.train_predictions = model.predict(data, offset=offset)
    return model


def _select_factor_variables(
    factor_table: pl.DataFrame,
    include_variables: list[str] | None,
    exclude_variables: list[str] | None,
) -> list[str]:
    """Select factor-table variables while preserving their sheet order."""
    all_variables = [
        variable
        for variable in factor_table["Variable"].unique()
        if variable != "intercept"
    ]
    if include_variables is not None:
        unknown = [
            variable for variable in include_variables if variable not in all_variables
        ]
        if unknown:
            raise ValueError(
                f"include_variables contains names not found in the sheet: {unknown}"
            )
        return [variable for variable in all_variables if variable in include_variables]
    if exclude_variables is not None:
        return [
            variable for variable in all_variables if variable not in exclude_variables
        ]
    return all_variables


def _read_factor_table(filepath: str | Path, sheet_name: str) -> pl.DataFrame:
    """Read and normalize an Excel factor table."""
    try:
        import openpyxl  # noqa: F401
    except ImportError as exc:
        raise ImportError(
            "openpyxl is required to read Excel files.\n"
            "Install it with:  pip install openpyxl"
        ) from exc

    factor_table = pl.read_excel(filepath, sheet_name=sheet_name, engine="openpyxl")
    missing_columns = {"Variable", "Level", "Factor"} - set(factor_table.columns)
    if missing_columns:
        raise ValueError(
            f"Excel sheet '{sheet_name}' is missing required columns: "
            f"{sorted(missing_columns)}.  Found: {factor_table.columns}"
        )
    return factor_table.with_columns(
        pl.col("Variable").cast(pl.String),
        pl.col("Level").cast(pl.String),
        pl.col("Factor").cast(pl.Float64),
    )


def _weighted_decile_levels(
    values: np.ndarray,
    weights: np.ndarray,
) -> np.ndarray:
    """Return the weighted decile labels used by AvE tables."""
    breaks = np.quantile(
        values,
        q=np.linspace(0, 1, 11),
        weights=weights,
        method="inverted_cdf",
    )
    breaks = sorted(set(dict.fromkeys(breaks)))[1:-1]
    return (
        pl.Series(values)
        .cut(
            breaks,
            labels=[str(index) for index in range(1, 11)],
            left_closed=True,
        )
        .cast(pl.String)
        .to_numpy()
    )


def _weighted_feat_map(
    Xt_df: pl.DataFrame,
    feats: list[str],
    w_arr: np.ndarray,
) -> dict[str, float]:
    """Sum of exposure weight per dummy feature column: {feat_name: total_weight}."""
    return {
        feat: float((Xt_df[feat].to_numpy() * w_arr).sum())
        for feat in feats
        if feat in Xt_df.columns
    }


def _make_row(
    var_col: str,
    level: str,
    weight: float,
    train_coef: float,
    fold_names: list[str],
    fold_coef_map: dict[str, dict[str, float]],
    feat: str | None,
    *,
    calib_weight: float | None = None,
) -> dict[str, Any]:
    """Build one row dict for the relativities table."""
    row: dict[str, Any] = {
        "variable": var_col,
        "level": level,
        "weight": weight,
        "train_coef": train_coef,
    }
    if calib_weight is not None:
        row["calib_weight"] = calib_weight
    for fn in fold_names:
        row[fn] = fold_coef_map[fn].get(feat, 0.0) if feat is not None else 0.0
    return row


def _validate_training_inputs(y, weights, offset) -> None:
    """Validate target, exposure, and offset arrays before model fitting."""
    y = y.to_numpy()
    if len(y) == 0 or not np.all(np.isfinite(y)):
        raise ValueError("target_col must contain at least one finite numeric value.")
    if weights is not None:
        weights = weights.to_numpy()
        if (
            not np.all(np.isfinite(weights))
            or np.any(weights < 0)
            or weights.sum() <= 0
        ):
            raise ValueError(
                "weight_col must contain finite non-negative values with a positive total."
            )
    if offset is not None:
        offset = offset.to_numpy()
        if offset is not None and (
            not np.all(np.isfinite(offset)) or len(offset) != len(y)
        ):
            raise ValueError(
                "offset_col must contain finite values aligned with target_col."
            )


class ModelingTool:
    """
    End-to-end elastic net GLM tool for insurance loss ratio modelling.

    All DataFrame arguments use **polars**.

    Workflow
    --------
    1. **Variable creation** — :meth:`add_variable` + :meth:`univariate_plot`.
    2. **Model fitting** — :meth:`fit_model` (CV or fixed alpha).
       :meth:`fit_cv_stability` evaluates coefficient stability across folds
       defined by a user-supplied column.
    3. **Variable evaluation** — :meth:`ae_chart`.
    4. **Model comparison** — :meth:`compare_models`.
    5. **Persistence** — :meth:`save` / :meth:`load`.

    Parameters
    ----------
    data : pl.DataFrame
        Source dataset.
    target_col : str
        Column name of the loss ratio target.
    weight_col : str, optional
        Column name of the exposure weights (e.g. earned premium).
    offset_col : str, optional
        Column name of a pre-existing linear predictor offset.  Values must
        be on the **log scale** (i.e. ``log(existing_prediction)``) for GLM
        versions fitted via :meth:`fit_model`.
    link : str or glum link, optional
        Default GLM link function applied to all :meth:`fit_model` calls.
    drop_reference : {'max_weight', 'first'}
        Which level to drop when one-hot encoding categorical variables.
        ``'max_weight'`` (default) drops the level with the highest total
        exposure weight.  ``'first'`` drops the first level alphabetically.
    cv_column : str, optional
        Column whose unique values define predefined CV folds for
        hyperparameter selection in :meth:`fit_model`.  Any hashable value is
        accepted as a fold label and is converted to a
        :class:`~sklearn.model_selection.PredefinedSplit` automatically.
        Pass an explicit ``cv=<int>`` to :meth:`fit_model` to override.
    current_version : str, optional
        Version name treated as active when no version is specified.
    base_version : str, optional
        Baseline version name used by default in comparison methods.
    """

    def __init__(
        self,
        data: pl.DataFrame,
        target_col: str,
        weight_col: str | None = None,
        offset_col: str | None = None,
        link: Any = None,
        tweedie_power: float = 1.5,
        drop_reference: str = "max_weight",
        cv_column: str | None = None,
        current_version: str | None = None,
        base_version: str | None = None,
    ):
        if not isinstance(data, pl.DataFrame):
            raise TypeError(
                f"data must be a polars DataFrame, got {type(data).__name__}."
            )
        if not isinstance(target_col, str) or target_col not in data.columns:
            raise ValueError(f"target_col '{target_col}' not found in data.")
        _validate_numeric_column(data, "weight_col", weight_col)
        _validate_numeric_column(data, "offset_col", offset_col)
        if not data[target_col].dtype.is_numeric():
            raise ValueError(
                f"target_col '{target_col}' must be numeric, got {data[target_col].dtype}."
            )
        if drop_reference not in {"max_weight", "first"}:
            raise ValueError("drop_reference must be 'max_weight' or 'first'.")
        if cv_column is not None and cv_column not in data.columns:
            raise ValueError(
                f"cv_column '{cv_column}' not found in data.  "
                f"Available columns: {data.columns}"
            )

        _validate_training_inputs(
            data[target_col],
            data[weight_col] if weight_col else None,
            data[offset_col] if offset_col else None,
        )

        self.data = data
        self.target_col = target_col
        self.weight_col = weight_col
        self.offset_col = offset_col
        self.drop_reference = drop_reference
        self.cv_column = cv_column
        self.link = link
        self.tweedie_power = tweedie_power
        self.current_version = current_version
        self.base_version = base_version
        self.variable_configs: dict[str, VariableConfig] = {}
        self.model_versions: dict[str, ModelVersion] = {}

    # ── Properties ───────────────────────────────────────────────────────────

    @property
    def _y(self) -> pl.Series:
        return self.data[self.target_col]

    @property
    def _y_array(self) -> np.ndarray:
        """Target values as a float64 numpy array."""
        return self._y.to_numpy().astype(float)

    @property
    def _weights(self) -> pl.Series | None:
        return self.data[self.weight_col] if self.weight_col else None

    @property
    def _weights_array(self) -> np.ndarray | None:
        """Exposure weights as a float64 numpy array, or None if no weight column."""
        return (
            self._weights.to_numpy().astype(float)
            if self._weights is not None
            else None
        )

    @property
    def _offset_array(self) -> np.ndarray | None:
        """Offset values as a float64 numpy array, or None if no offset column."""
        return (
            self.data[self.offset_col].to_numpy().astype(float)
            if self.offset_col is not None
            else None
        )

    # ── Variable management ───────────────────────────────────────────────────

    def add_variable(
        self,
        col: str,
        config: VariableConfig | None = None,
        input_cols: list[str] | None = None,
        custom_transform: Callable | None = None,
        **kwargs,
    ) -> ModelingTool:
        """
        Register preprocessing for a variable.

        Three calling styles:

        1. **Pass a VariableConfig directly**::

               tool.add_variable('x', config=VariableConfig(col='x', n_bins=10))

        2. **Keyword arguments** (most common)::

               tool.add_variable('vehicle_age', cap_upper=0.99, log_transform=True)
               tool.add_variable('state', encoding='onehot')
               tool.add_variable('driver_age', bin_edges=[16,25,35,50,65,100])

        3. **Named built-in transformation** (new name for one scalar pipeline)::

               tool.add_variable(
                   'vehicle_value_logged',
                   input_cols=['vehicle_value'],
                   log_transform=True,
               )

        4. **Custom derived variable** (new named variable from one or more columns)::

               tool.add_variable(
                   'age_x_veh',
                   input_cols=['driver_age', 'vehicle_age'],
                   custom_transform=lambda df: df['driver_age'] * df['vehicle_age'],
                   cap_upper=0.99,
               )

        For categorical variables, ``custom_transform`` is applied before encoding::

               tool.add_variable(
                   'state',
                   custom_transform=lambda df: [
                       'South' if v in ('TX', 'FL') else 'Other'
                       for v in df['state']
                   ],
                   encoding='onehot',
               )

        Parameters
        ----------
        col : str
            Output variable name (must match ``config.col`` when passing a
            config directly, and must exist in ``data`` for dtype detection
            when no options are given).
        config : VariableConfig, optional
            A fully constructed config.  Mutually exclusive with keyword args.
        input_cols : list of str, optional
            Source columns for a derived variable. Exactly one source without
            ``custom_transform`` creates a named scalar built-in pipeline.
            One or more sources may be paired with ``custom_transform``.
        custom_transform : callable, optional
            Function called once with a DataFrame containing ``input_cols``
            (or ``col`` for a single-column transform). Derived inputs are
            resolved recursively. Pipeline aliases supply fitted scalar values;
            custom-derived inputs supply raw, pre-pipeline values.
        **kwargs
            Any :class:`~modeling_tool.variable.VariableConfig` field
            (e.g. ``n_bins``, ``bin_edges`` / ``breakpoints``, ``cap_upper``,
            ``log_transform``, ``encoding``, ``standardize``).

        If no arguments are provided, a default config is inferred from dtype.
        """
        if config is not None:
            if not isinstance(config, VariableConfig):
                raise TypeError("config must be a VariableConfig instance.")
            if config.col != col:
                raise ValueError("config.col must match the add_variable output name.")
            if kwargs or input_cols is not None or custom_transform is not None:
                raise ValueError(
                    "Pass either config or individual variable options, not both."
                )
            self.variable_configs[col] = config
            return self

        self.variable_configs[col] = _create_variable_config(
            self.data,
            col,
            input_cols,
            custom_transform,
            kwargs,
        )
        return self

    def get_variable_config(self, col: str) -> VariableConfig | None:
        """Return the registered :class:`VariableConfig` for *col*."""
        return self.variable_configs.get(col)

    def list_variables(self) -> pl.DataFrame:
        """Summary table of all registered variable configs."""
        rows = []
        for col, cfg in self.variable_configs.items():
            rows.append(
                {
                    "col": col,
                    "input_cols": str(cfg.input_cols) if cfg.input_cols else None,
                    "is_categorical": cfg.is_categorical,
                    "cap_lower": cfg.cap_lower,
                    "cap_upper": cfg.cap_upper,
                    "log_transform": cfg.log_transform,
                    "n_bins": cfg.n_bins,
                    "bin_edges": str(cfg.bin_edges) if cfg.bin_edges else None,
                    "standardize": cfg.standardize,
                    "encoding": cfg.encoding,
                    "impute_strategy": cfg.impute_strategy,
                    "custom_transform": cfg.custom_transform is not None,
                }
            )
        return pl.DataFrame(rows) if rows else pl.DataFrame()

    # ── Exploration ───────────────────────────────────────────────────────────

    def univariate_plot(
        self,
        col: str,
        n_bins: int = 10,
        breaks: list[float] | None = None,
        figsize: tuple[int, int] | None = None,
        version: str | None = None,
        show: bool = True,
    ) -> plt.Figure:
        """
        Weighted mean target vs each level (or quantile bin) of *col*.

        For continuous variables, bins into ``n_bins`` quantile groups.
        The sentinel value ``-999999999`` is labelled ``'Missing'``.

        Parameters
        ----------
        col : str
            Column to analyse.  Does not need to be in ``variable_configs``.
        n_bins : int
            Number of quantile bins for continuous variables (default 10).
        breaks : list of float, optional
            Explicit bin edges for continuous variables.  Overrides ``n_bins``.
        figsize : tuple of int, optional
            Figure size ``(width, height)`` in inches.
        version : str, optional
            When supplied, the fitted preprocessor from this version is used
            to resolve bin labels for *col* (consistent with
            :meth:`summary_table`).
        """
        preprocessor = None
        if version is not None:
            mv = self._get_version(version)
            preprocessor = getattr(mv, "preprocessor", None)
            if col in self.variable_configs and (
                preprocessor is None or col not in preprocessor._params
            ):
                preprocessor = _fit_plot_preprocessor(
                    col,
                    self.data,
                    self.variable_configs,
                    self._weights_array,
                )
        elif col in self.variable_configs:
            preprocessor = _fit_plot_preprocessor(
                col,
                self.data,
                self.variable_configs,
                self._weights_array,
            )

        fig = univariate_plot(
            self.data,
            self._y,
            col=col,
            weights=self._weights,
            n_bins=n_bins,
            breaks=breaks,
            figsize=figsize,
            preprocessor=preprocessor,
        )

        if show:
            plt.show()
        return fig

    # ── Bin suggestion ────────────────────────────────────────────────────────

    def _data_with_derived_col(self, col: str) -> pl.DataFrame:
        """Return data guaranteed to contain *col*, deriving it via config if needed."""
        cfg = self.variable_configs.get(col)
        is_derived = cfg is not None and (
            cfg.input_cols is not None or cfg.custom_transform is not None
        )
        if col in self.data.columns and not is_derived:
            return self.data
        if cfg is None or not is_derived:
            raise ValueError(
                f"Column '{col}' not found in data and has no derivable config "
                "(register a pipeline alias or custom derived variable)."
            )
        stripped = self._strip_config(cfg)
        configs = dict(self.variable_configs)
        configs[col] = stripped
        prep = _build_preprocessor([col], self.data, configs)
        prep.fit(self.data, weights=self._weights_array)
        return self.data.with_columns(prep.transform(self.data)[col])

    def suggest_bins_quantile(
        self,
        col: str,
        n_bins: int = 10,
        verbose: bool = True,
        **kwargs,
    ) -> list[float]:
        """Equal-weight quantile breakpoints for *col*. Shortcut for ``suggest_bins(methods=['quantile'])[...]``."""
        from .bin_suggestor import suggest_bins_quantile as _fn

        return _fn(
            col,
            self._data_with_derived_col(col),
            n_bins=n_bins,
            weights=self._weights,
            verbose=verbose,
            **kwargs,
        )

    def suggest_bins_equal_width(
        self,
        col: str,
        n_bins: int = 10,
        verbose: bool = True,
        **kwargs,
    ) -> list[float]:
        """Equal-width breakpoints for *col*. Shortcut for ``suggest_bins(methods=['equal_width'])[...]``."""
        from .bin_suggestor import suggest_bins_equal_width as _fn

        return _fn(
            col,
            self._data_with_derived_col(col),
            n_bins=n_bins,
            verbose=verbose,
            **kwargs,
        )

    def suggest_bins_gbm(
        self,
        col: str,
        max_splits: int = 20,
        verbose: bool = True,
        **kwargs,
    ) -> list[float]:
        """GBM-derived breakpoints for *col*. Shortcut for ``suggest_bins(methods=['gbm'])[...]``."""
        from .bin_suggestor import suggest_bins_gbm as _fn

        return _fn(
            col,
            self._data_with_derived_col(col),
            self._y,
            weights=self._weights,
            max_splits=max_splits,
            verbose=verbose,
            **kwargs,
        )

    def suggest_bins_optbin(
        self,
        col: str,
        max_n_bins: int = 10,
        monotonic_trend: str = "auto",
        verbose: bool = True,
        **kwargs,
    ) -> list[float]:
        """Optimal binning breakpoints for *col*. Shortcut for ``suggest_bins(methods=['optbin'])[...]``."""
        from .bin_suggestor import suggest_bins_optbin as _fn

        return _fn(
            col,
            self._data_with_derived_col(col),
            self._y,
            weights=self._weights,
            max_n_bins=max_n_bins,
            monotonic_trend=monotonic_trend,
            verbose=verbose,
            **kwargs,
        )

    def suggest_bins(
        self,
        col: str,
        methods: Sequence[str] = ("quantile", "equal_width", "optbin", "gbm"),
        n_bins: int = 10,
        max_splits: int = 20,
        show_plot: bool = False,
        figsize: tuple[int, int] | None = None,
        **method_kwargs: Any,
    ) -> dict[str, list[float]]:
        """
        Run multiple bin-suggestion strategies for a continuous variable.

        Prints each method's splits, then shows a weighted histogram with all
        split points overlaid as colour-coded vertical lines so results can be
        compared visually before committing to any breakpoints.

        Parameters
        ----------
        col : str
            Continuous variable to analyse.  Does not need to be in the model.
        methods : sequence of str
            Any subset of ``"quantile"``, ``"equal_width"``, ``"optbin"``,
            ``"gbm"``.  Defaults to running all four.
        n_bins : int
            Target bin count for ``"quantile"`` and ``"equal_width"``.
        max_splits : int
            Maximum thresholds returned by the ``"gbm"`` method (selected by
            frequency of use across all trees).
        show_plot : bool
            Display the distribution chart after all methods run.
        figsize : tuple of int, optional
            Figure size ``(width, height)`` in inches for the distribution chart.
        method_kwargs
            Forward kwargs to individual methods via ``quantile_kwargs``,
            ``equal_width_kwargs``, ``optbin_kwargs``, or ``gbm_kwargs``
            as dicts.

        Returns
        -------
        dict[str, list[float]]
            Method name → sorted list of suggested split points.

        Examples
        --------
        >>> splits = tool.suggest_bins('driver_age')
        >>> splits = tool.suggest_bins(
        ...     'vehicle_value',
        ...     methods=['optbin', 'gbm'],
        ...     optbin_kwargs={'max_n_bins': 6, 'monotonic_trend': 'auto'},
        ...     gbm_kwargs={'n_estimators': 200, 'learning_rate': 0.05},
        ... )
        >>> # Apply the optbin result directly
        >>> tool.add_variable('vehicle_value', breakpoints=splits['optbin'])
        """
        from .bin_suggestor import suggest_bins as _suggest_bins

        return _suggest_bins(
            col=col,
            X=self._data_with_derived_col(col),
            y=self._y,
            weights=self._weights,
            methods=methods,
            n_bins=n_bins,
            max_splits=max_splits,
            figsize=figsize,
            **method_kwargs,
        )

    # ── Model fitting ─────────────────────────────────────────────────────────

    def _resolve_cv(
        self,
        alpha: float | None,
        cv: int | None,
        use_cv: bool,
    ) -> tuple[bool, Any]:
        """Resolve fixed-alpha and predefined cross-validation behavior."""
        if alpha is not None:
            return False, None
        if cv is not None:
            return use_cv, cv
        if self.cv_column is None:
            return use_cv, 5

        from sklearn.model_selection import PredefinedSplit

        fold_values = self.data[self.cv_column]
        unique_folds = sorted(set(fold_values))
        fold_map = {fold: index for index, fold in enumerate(unique_folds)}
        test_fold = [fold_map[fold] for fold in fold_values]
        return use_cv, PredefinedSplit(test_fold)

    def fit_model(
        self,
        variables: list[str],
        version: str,
        alpha: float | None = None,
        l1_ratio: float | list[float] = 0.5,
        use_cv: bool = True,
        cv: int | None = None,
        family: Any = None,
        link: str | None = None,
        tweedie_power: float | None = 1.50,
        preprocessor: Preprocessor | None = None,
        max_iter: int = 1000,
        gradient_tol: float | None = None,
        fit_intercept: bool = True,
        print_summary: bool = True,
        n_jobs: int | None = None,
        alphas: np.ndarray | None = None,
    ) -> ModelVersion:
        """
        Fit an elastic net GLM and store it as a named version.

        Variables without a registered :class:`VariableConfig` are given
        sensible defaults based on their dtype.

        Parameters
        ----------
        variables : list of str
            Predictor column names.  Multi-input derived variable names can
            appear here if their config has been registered via
            :meth:`add_variable`.
        version : str
            Version label (e.g. ``'v1'``, ``'with_geo'``).
        alpha : float, optional
            Fixed regularisation strength.  Ignored when ``use_cv=True``.
            Pass ``0.0`` for an unpenalised GLM.
        l1_ratio : float or list of float
            Elastic-net mixing (0=ridge, 1=lasso).  A list triggers a CV
            grid search.
        use_cv : bool
            Select best alpha (and l1_ratio if a list) via CV.
        cv : int, optional
            Number of k-fold splits for hyperparameter selection.  When
            ``None`` (default) and a ``cv_column`` was supplied at
            construction, a :class:`~sklearn.model_selection.PredefinedSplit`
            is built from that column automatically.  When ``None`` and no
            ``cv_column`` exists, falls back to 5-fold CV.  Pass an explicit
            integer to override the ``cv_column`` for this specific fit.
        family : glum distribution, optional
            GLM response distribution.  Defaults to
            ``TweedieDistribution(power=tweedie_power)``.
        link : str or glum link, optional
            GLM link function.  Overrides the tool-level default for this fit.
        tweedie_power : float, optional
            Tweedie variance power (default 1.5).  Ignored when *family* is
            supplied explicitly.
        preprocessor : Preprocessor, optional
            Pre-built preprocessor to use instead of constructing one from
            ``variable_configs``.  Useful for reusing an existing transform.
        max_iter : int
            Maximum IRLS iterations (default 1000).
        gradient_tol : float, optional
            Convergence tolerance on the gradient.  Uses glum's default when
            ``None``.
        fit_intercept : bool
            Whether to fit an intercept term (default ``True``).
        print_summary : bool
            Print the coefficient summary table after fitting (default ``True``).
        n_jobs : int, optional
            Number of parallel jobs for CV.  Defaults to ``None`` (1 job).
        """
        _validate_fit_request(variables, version)
        use_cv, resolved_cv = self._resolve_cv(alpha, cv, use_cv)
        if type(resolved_cv).__name__ == "PredefinedSplit":
            print("Using PredefinedSplit from cv_column.")
        mv = fit_model(
            X=self.data,
            y=self._y_array,
            variables=variables,
            version_name=version,
            configs=self.variable_configs,
            weights=self._weights_array,
            offset=self._offset_array,
            family=family,
            link=link or self.link,
            tweedie_power=self.tweedie_power
            if tweedie_power is None
            else tweedie_power,
            preprocessor=preprocessor,
            alpha=alpha,
            l1_ratio=l1_ratio,
            use_cv=use_cv,
            cv=resolved_cv,
            max_iter=max_iter,
            gradient_tol=gradient_tol,
            fit_intercept=fit_intercept,
            drop_reference=self.drop_reference,
            n_jobs=n_jobs,
            alphas=alphas,
        )
        self.model_versions[version] = mv
        self.current_version = version

        if print_summary:
            self.model_summary(version)
        return None

    def fit_cv_stability(
        self,
        version: str | None = None,
        family: Any = None,
        link: str | None = None,
        tweedie_power: float | None = None,
        plot: bool = True,
        show: bool = True,
    ) -> pl.DataFrame:
        """
        Assess coefficient stability using user-defined CV folds.

        For each unique value in ``self.cv_column``, the model is trained on all
        other rows and the coefficients are stored.  The geometric mean,
        standard deviation, and coefficient of variation (%) across folds
        are appended as summary rows.

        Parameters
        ----------
        version : str, optional
            Borrow ``alpha`` and ``l1_ratio`` from a previously fitted version.
            Uses current version when ``None``.
        family : glum distribution, optional
            Override the GLM family used for fold refits.
        link : str or glum link, optional
            Override the GLM link used for fold refits.
        tweedie_power : float, optional
            Override the Tweedie variance power used for fold refits.
        plot : bool
            Show a coefficient stability box-plot after fitting.
        show : bool
            Call ``plt.show()`` after plotting.

        Returns
        -------
        pl.DataFrame
            Rows = one per fold + ``'geomean'``, ``'std'``, ``'cv_pct'``.
            Columns = ``'fold'`` + intercept + one per feature.
        """
        if version is not None and version in self.model_versions:
            mv = self.model_versions[version]
        elif version is None:
            print(
                "Using current version since none was specified or specified version does not exist."
            )
            mv = self.model_versions[self.current_version]
        else:
            raise AttributeError(
                "No model has yet been fit, please fit a model first and then specify a version."
            )

        resolved_alpha = mv.alpha
        resolved_l1 = mv.l1_ratio

        stability = fit_cv_stability(
            X=self.data,
            y=self._y_array,
            variables=mv.variables,
            configs=self.variable_configs,
            fold_col=self.cv_column,
            weights=self._weights_array,
            offset=self._offset_array,
            family=family or mv.family,
            link=link or self.link,
            tweedie_power=tweedie_power or mv.tweedie_power,
            alpha=resolved_alpha if resolved_alpha is not None else 0.01,
            l1_ratio=resolved_l1 if resolved_l1 is not None else 0.5,
            drop_reference=self.drop_reference,
            gradient_tol=mv.gradient_tol if hasattr(mv, "gradient_tol") else None,
            preprocessor=mv.preprocessor,
        )

        if plot:
            _fig = cv_stability_plot(stability)
            if show:
                plt.show()

        self.model_versions[mv.name].cv_stability = stability

        return stability

    def set_base_version(self, version: str) -> ModelingTool:
        """
        Set the baseline version used by default in comparison methods.

        Parameters
        ----------
        version : str
            Registered model version name, or a column name in ``data``
            containing pre-computed baseline predictions.
        """
        if version not in self.model_versions and version not in self.data.columns:
            raise ValueError(
                f"Version '{version}' not found and not in data columns. Available versions: {list(self.model_versions.keys())}"
            )

        self.base_version = version
        return self

    def predict(
        self,
        data: pl.DataFrame,
        version: str | None = None,
        missing_factor: float = 1.0,
        offset: np.ndarray | None = None,
    ) -> np.ndarray:
        """Generate predictions for *data* using the specified model version.

        Parameters
        ----------
        data : pl.DataFrame
            Dataset to score.  Must contain all columns used by the version.
        version : str, optional
            Version to use for predictions.  Defaults to ``current_version``.
        missing_factor : float
            Multiplicative factor applied to rows whose level is absent from
            a factor-table model (default 1.0).  Ignored for fitted GLM
            versions.
        offset : np.ndarray, optional
            One-dimensional array of offset values aligned with *data*.  When
            ``None``, the tool-level offset column is used if one was set.
            Scale depends on the version type: **log scale** (linear predictor)
            for fitted GLM versions; **response scale** (multiplicative factor)
            for Excel factor-table versions.

        Returns
        -------
        np.ndarray
            Predicted values, one per row in *data*.
        """
        if not isinstance(data, pl.DataFrame):
            raise TypeError("data must be a polars DataFrame.")
        if not isinstance(missing_factor, (int, float)) or not np.isfinite(
            missing_factor
        ):
            raise ValueError("missing_factor must be finite.")
        mv = self._get_version(version or self.current_version)
        if offset is None:
            if self.offset_col is not None:
                if self.offset_col not in data.columns:
                    raise ValueError(
                        f"offset column '{self.offset_col}' not found in scoring data."
                    )
                offset = data[self.offset_col].to_numpy().astype(float)
        if offset is not None and (
            np.asarray(offset).ndim != 1
            or len(offset) != len(data)
            or not np.all(np.isfinite(offset))
        ):
            raise ValueError(
                "offset must be a finite one-dimensional array aligned with data."
            )

        if isinstance(mv, FactorModelVersion):
            if (
                mv.offset_col
                and mv.offset_col in data.columns
                and mv.offset_col != self.offset_col
            ):
                offset = data[mv.offset_col].to_numpy().astype(float)
            return mv.predict(data, missing_factor=missing_factor, offset=offset)
        else:
            return mv.predict(data, offset=offset)

    # ── Excel factor version ──────────────────────────────────────────────────

    def _excel_preprocessor(
        self,
        variables: list[str],
        base_version: str | None,
    ) -> Preprocessor | None:
        """Resolve the preprocessor used by an Excel factor model."""
        if base_version is not None:
            base_model = self._get_version(base_version)
            return getattr(base_model, "preprocessor", None)
        preprocessor = _build_preprocessor(
            variables,
            self.data,
            self.variable_configs,
        )
        preprocessor.fit(self.data, self._y_array, weights=self._weights_array)
        return preprocessor

    def _validate_direct_factor_variables(
        self,
        direct_variables: list[str],
    ) -> None:
        """Ensure direct-lookup factors exist in non-empty training data."""
        if not len(self.data):
            return
        missing = [
            variable
            for variable in direct_variables
            if variable not in self.data.columns
        ]
        if missing:
            raise ValueError(
                f"Variables {missing} are not covered by any preprocessor "
                "and are not found in the data columns. "
                "Ensure these columns exist or specify base_version."
            )

    def _initialize_factor_predictions(
        self,
        model: FactorModelVersion,
        missing_factor: float,
    ) -> None:
        """Populate in-sample predictions for an Excel factor model."""
        if not len(self.data):
            model.train_predictions = np.array([])
            return
        offset = None
        if model.offset_col and model.offset_col in self.data.columns:
            offset = self.data[model.offset_col].to_numpy().astype(float)
        model.train_predictions = model.predict(
            self.data,
            missing_factor=missing_factor,
            offset=offset,
        )

    def add_excel_version(
        self,
        filepath: str | Path,
        sheet_name: str,
        version: str = "excel",
        missing_factor: float = 1.0,
        base_version: str | None = None,
        offset_col: str | None = None,
        include_variables: list[str] | None = None,
        exclude_variables: list[str] | None = None,
    ) -> ModelingTool:
        """
        Load factors from an Excel sheet and register them as a new model version.

        The sheet must have columns **Variable**, **Level**, **Factor**.
        Level strings for variables covered by a fitted preprocessor must match
        :meth:`summary_table` output (e.g. ``'TX'``, ``'[16, 25) (base)'``,
        ``'Missing'``).  For all other variables the raw column value is used as
        the level (direct string match).

        An optional row with ``Variable='intercept'`` and ``Level='intercept'``
        applies a global multiplicative factor to every prediction.

        Parameters
        ----------
        filepath : str
            Path to the ``.xlsx`` workbook.
        sheet_name : str
            Sheet name containing the Variable / Level / Factor table.
        version : str
            Version label to register (default ``'excel'``).
        missing_factor : float
            Factor applied to rows whose level is absent from the table
            (default 1.0, with a printed warning).
        base_version : str, optional
            Name of an existing fitted version whose preprocessor is used for
            numeric/binned level resolution.  When ``None``, the preprocessor
            that covers the most Excel variables is chosen automatically.
            Variables not covered by any preprocessor fall back to direct
            string lookup.
        offset_col : str, optional
            Column in the tool's data containing a per-row multiplicative
            offset on the **response scale** (i.e. already exponentiated).
            The factor product is multiplied element-wise by this column's
            values.
        include_variables : list of str, optional
            If provided, only variables in this list are loaded from the sheet.
            Mutually exclusive with ``exclude_variables``.
        exclude_variables : list of str, optional
            Variables to drop from the sheet before loading.  Mutually
            exclusive with ``include_variables``.

        Returns
        -------
        ModelingTool
            Returns ``self`` to support method chaining.
        """
        if include_variables is not None and exclude_variables is not None:
            raise ValueError(
                "Specify only one of include_variables or exclude_variables, not both."
            )

        factor_table = _read_factor_table(filepath, sheet_name)
        variables = _select_factor_variables(
            factor_table,
            include_variables,
            exclude_variables,
        )
        factor_table = factor_table.filter(
            pl.col("Variable").is_in(variables) | (pl.col("Variable") == "intercept")
        )
        prep = self._excel_preprocessor(variables, base_version)
        preprocessor_vars = (
            [variable for variable in variables if variable in prep.configs]
            if prep is not None
            else []
        )
        direct_vars = [
            variable for variable in variables if variable not in preprocessor_vars
        ]
        self._validate_direct_factor_variables(direct_vars)
        fmv = FactorModelVersion(
            name=version,
            variables=variables,
            factor_table=factor_table,
            preprocessor=prep,
            preprocessor_vars=preprocessor_vars,
            train_predictions=np.array([]),
            offset_col=offset_col,
        )
        self._initialize_factor_predictions(fmv, missing_factor)
        self.model_versions[version] = fmv
        print(
            f"  [Excel] Version '{version}' registered — "
            f"{len(variables)} variable(s): "
            f"{len(preprocessor_vars)} preprocessor-resolved, "
            f"{len(direct_vars)} direct-lookup."
        )
        return self

    # ── Model summary ─────────────────────────────────────────────────────────

    def model_summary(self, version: str | None = None) -> pl.DataFrame:
        """Print and return the coefficient table for *version*.

        Parameters
        ----------
        version : str, optional
            Model version to summarise.  Uses current version if ``None``.
        """
        mv = self._get_version(version or self.current_version)
        if isinstance(mv, FactorModelVersion):
            raise TypeError(
                f"Version '{version}' is an Excel factor model; "
                "model_summary is not applicable."
            )
        nonzero = (
            int((mv.coefficients["coefficient"] != 0).sum()) - 1
        )  # exclude intercept

        top_data = [
            ("Model Version:", mv.name),
            ("Dep. Variable:", self.target_col),
            (
                "Model:",
                f"GLM, Penalty Weight = {np.round(mv.alpha, 3)}, L1 Ratio = {np.round(mv.l1_ratio, 3)}",
            ),
            ("Model Family:", mv.family.__class__.__name__),
            ("Link Function:", mv.link.__class__.__name__),
            ("Offset Column:", self.offset_col),
            ("Method:", "IRLS-CD"),
            ("Fit Date:", mv.fit_info.get("Fit_Time", "N/A")),
            ("No. Iterations:", len(mv.glm.diagnostics_)),
            ("No. Observations:", len(self.data)),
            ("# Features:", len(mv.feature_names)),
            ("# Nonzero Coefs:", nonzero),
        ]
        top_part = tabulate(top_data, tablefmt="plain")

        coef_table = tabulate(
            mv.coefficients.with_columns(pl.col("coefficient").exp()).to_dict()
        )

        print("=" * 80)
        print(top_part)
        print("=" * 80)
        print(coef_table)

        return None

    def coefficient_plot(
        self,
        version: str | None = None,
        top_n: int = 30,
        figsize: tuple[int, int] | None = None,
        show: bool = True,
    ) -> plt.Figure:
        """Horizontal bar chart of coefficients for *version*.

        Parameters
        ----------
        version : str, optional
            Model version to plot.  Uses current version if ``None``.
        top_n : int
            Maximum number of features to show (default 30).
        figsize : tuple of int, optional
            Figure size ``(width, height)`` in inches.
        show : bool
            Call ``plt.show()`` after rendering.
        """
        mv = self._get_version(version or self.current_version)
        if isinstance(mv, FactorModelVersion):
            raise TypeError(
                f"Version '{version}' is an Excel factor model; "
                "coefficient_plot is not applicable."
            )
        fig = coefficient_plot(
            mv.coefficients, version_name=version, top_n=top_n, figsize=figsize
        )
        if show:
            plt.show()
        return fig

    # ── Private helpers for relativities_table ───────────────────────────────

    def _get_fold_info(
        self, mv: ModelVersion
    ) -> tuple[list[str], dict[str, dict[str, float]]]:
        """Refit on each fold; return (fold_names, fold_label -> {feat: coef})."""
        if mv.cv_stability is None:
            stability = fit_cv_stability(
                X=self.data,
                y=self._y_array,
                variables=mv.variables,
                configs=self.variable_configs,
                fold_col=self.cv_column,
                weights=self._weights_array,
                offset=self._offset_array,
                family=mv.family,
                link=mv.link,
                alpha=mv.alpha,
                l1_ratio=mv.l1_ratio,
                drop_reference=self.drop_reference,
                preprocessor=mv.preprocessor,
            )
            mv.cv_stability = stability
        else:
            stability = mv.cv_stability
        fold_rows = stability.filter(
            ~pl.col("fold").is_in(["geomean", "std", "cv_pct"])
        )
        raw_fold_names = fold_rows["fold"].to_list()
        fold_names = [f"fold_{name}" for name in raw_fold_names]
        fold_coef_map: dict[str, dict[str, float]] = {
            display_name: {k: v for k, v in row_d.items() if k != "fold"}
            for display_name, row_d in zip(fold_names, fold_rows.to_dicts())
        }
        return fold_names, fold_coef_map

    def _get_calib_arrays(
        self,
        prep: Any,
        calib_df: pl.DataFrame | None,
    ) -> tuple[pl.DataFrame | None, np.ndarray | None, float]:
        """Transform calib_df; return (Xt_calib, w_calib, total_calib_w)."""
        if calib_df is None:
            return None, None, 0.0
        # strict=False: unseen categorical levels (e.g. future EVALUATION_YEAR values) and
        # missing un-imputed numerics in calib are silently zeroed rather than raising.
        # Calib data here is used only for exposure-weight counting, not prediction.
        Xt_calib = prep.transform(calib_df, strict=False)
        w_calib = (
            calib_df[self.weight_col].to_numpy().astype(float)
            if self.weight_col and self.weight_col in calib_df.columns
            else np.ones(len(calib_df))
        )
        return Xt_calib, w_calib, float(w_calib.sum())

    def _cat_var_rows(
        self,
        var_col: str,
        p: FittedCategoricalParams,
        Xt_df: pl.DataFrame,
        w_arr: np.ndarray,
        total_w: float,
        coef_map: dict[str, float],
        fold_names: list[str],
        fold_coef_map: dict[str, dict[str, float]],
        Xt_calib: pl.DataFrame | None,
        w_calib: np.ndarray | None,
        total_calib_w: float,
    ) -> list[dict[str, Any]]:
        """Row dicts for one one-hot categorical variable."""
        categories = p.categories
        dropped = p.dropped_category
        feats = [f"{var_col}_{cat}" for cat in categories]

        train_fw = _weighted_feat_map(Xt_df, feats, w_arr)
        other_w = sum(train_fw.values())

        calib_fw = (
            _weighted_feat_map(Xt_calib, feats, w_calib) if Xt_calib is not None else {}
        )
        base_cw: float | None = (
            total_calib_w - sum(calib_fw.values()) if Xt_calib is not None else None
        )

        rows: list[dict[str, Any]] = []
        if dropped is not None:
            rows.append(
                _make_row(
                    var_col,
                    f"{dropped} (base)",
                    total_w - other_w,
                    0.0,
                    fold_names,
                    fold_coef_map,
                    None,
                    calib_weight=base_cw,
                )
            )
        for cat, feat in zip(categories, feats):
            rows.append(
                _make_row(
                    var_col,
                    str(cat),
                    train_fw.get(feat, 0.0),
                    coef_map.get(feat, 0.0),
                    fold_names,
                    fold_coef_map,
                    feat,
                    calib_weight=calib_fw.get(feat, 0.0)
                    if Xt_calib is not None
                    else None,
                )
            )
        return rows

    def _binned_var_rows(
        self,
        var_col: str,
        p: FittedBinnedNumericParams,
        cfg: Any,
        Xt_df: pl.DataFrame,
        w_arr: np.ndarray,
        total_w: float,
        coef_map: dict[str, float],
        fold_names: list[str],
        fold_coef_map: dict[str, dict[str, float]],
        Xt_calib: pl.DataFrame | None,
        w_calib: np.ndarray | None,
        total_calib_w: float,
    ) -> list[dict[str, Any]]:
        """Row dicts for one binned numeric variable."""
        _edges = p.bin_edges
        dropped_bin = p.dropped_bin
        all_labels = p.bin_labels

        missing_feat = f"{var_col}_missing"
        bin_feats = [
            f"{var_col}_{label}"
            for i, label in enumerate(all_labels)
            if i != dropped_bin
        ]
        all_feats = (
            [missing_feat] if missing_feat in Xt_df.columns else []
        ) + bin_feats

        train_fw = _weighted_feat_map(Xt_df, all_feats, w_arr)
        other_w = sum(train_fw.values())

        calib_fw = (
            _weighted_feat_map(Xt_calib, all_feats, w_calib)
            if Xt_calib is not None
            else {}
        )
        base_cw: float | None = (
            total_calib_w - sum(calib_fw.values()) if Xt_calib is not None else None
        )

        base_label = all_labels[dropped_bin]
        rows: list[dict[str, Any]] = [
            _make_row(
                var_col,
                f"{base_label} (base)",
                total_w - other_w,
                0.0,
                fold_names,
                fold_coef_map,
                None,
                calib_weight=base_cw,
            )
        ]
        if missing_feat in Xt_df.columns:
            rows.append(
                _make_row(
                    var_col,
                    "Missing",
                    train_fw.get(missing_feat, 0.0),
                    coef_map.get(missing_feat, 0.0),
                    fold_names,
                    fold_coef_map,
                    missing_feat,
                    calib_weight=calib_fw.get(missing_feat, 0.0)
                    if Xt_calib is not None
                    else None,
                )
            )
        for i, label in enumerate(all_labels):
            if i == dropped_bin:
                continue
            feat = f"{var_col}_{label}"
            rows.append(
                _make_row(
                    var_col,
                    label,
                    train_fw.get(feat, 0.0),
                    coef_map.get(feat, 0.0),
                    fold_names,
                    fold_coef_map,
                    feat,
                    calib_weight=calib_fw.get(feat, 0.0)
                    if Xt_calib is not None
                    else None,
                )
            )
        return rows

    def _encoded_variable_rows(
        self,
        variable: str,
        params: Any,
        config: VariableConfig,
        design: pl.DataFrame,
        weights: np.ndarray,
        total_weight: float,
        coefficients: dict[str, float],
        fold_names: list[str],
        fold_coefficients: dict[str, dict[str, float]],
        calibration_design: pl.DataFrame | None,
        calibration_weights: np.ndarray | None,
        total_calibration_weight: float,
    ) -> list[dict[str, Any]] | None:
        """Build summary rows for one encoded variable."""
        shared = (
            variable,
            params,
            design,
            weights,
            total_weight,
            coefficients,
            fold_names,
            fold_coefficients,
            calibration_design,
            calibration_weights,
            total_calibration_weight,
        )
        if isinstance(params, FittedCategoricalParams) and params.encoding == "onehot":
            return self._cat_var_rows(*shared)
        if isinstance(params, FittedBinnedNumericParams):
            return self._binned_var_rows(
                variable,
                params,
                config,
                *shared[2:],
            )
        return None

    def _model_summary_rows(
        self,
        model: ModelVersion,
        design: pl.DataFrame,
        weights: np.ndarray,
        total_weight: float,
        coefficients: dict[str, float],
        fold_names: list[str],
        fold_coefficients: dict[str, dict[str, float]],
        calibration_design: pl.DataFrame | None,
        calibration_weights: np.ndarray | None,
        total_calibration_weight: float,
    ) -> list[dict[str, Any]]:
        """Build summary rows for the model's discrete variables."""
        rows: list[dict[str, Any]] = []
        preprocessor = model.preprocessor
        for variable in model.variables:
            if variable not in preprocessor.configs:
                continue
            variable_rows = self._encoded_variable_rows(
                variable,
                preprocessor._params.get(variable),
                preprocessor.configs[variable],
                design,
                weights,
                total_weight,
                coefficients,
                fold_names,
                fold_coefficients,
                calibration_design,
                calibration_weights,
                total_calibration_weight,
            )
            rows.extend(variable_rows or [])
        return rows

    def _raw_extra_variable_rows(
        self,
        variable: str,
        weights: np.ndarray,
        calibration_data: pl.DataFrame | None,
        calibration_weights: np.ndarray | None,
    ) -> list[dict[str, Any]]:
        """Build zero-coefficient summary rows from raw resolved levels."""
        levels = _resolve_level(variable, self.data, preprocessor=None, n_bins=10)
        level_weights = (
            pl.DataFrame({"_level": levels, "_w": weights})
            .group_by("_level")
            .agg(pl.col("_w").sum())
        )
        calibration_level_weights: dict[str, float] = {}
        if (
            calibration_data is not None
            and variable in calibration_data.columns
            and calibration_weights is not None
        ):
            calibration_levels = _resolve_level(
                variable,
                calibration_data,
                preprocessor=None,
                n_bins=10,
            )
            grouped = (
                pl.DataFrame({"_level": calibration_levels, "_w": calibration_weights})
                .group_by("_level")
                .agg(pl.col("_w").sum())
            )
            calibration_level_weights = {
                row["_level"]: row["_w"] for row in grouped.iter_rows(named=True)
            }

        rows = []
        for label in _sort_labels(level_weights["_level"].to_list()):
            weight = float(level_weights.filter(pl.col("_level") == label)["_w"][0])
            calibration_weight = (
                calibration_level_weights.get(label, 0.0)
                if calibration_data is not None
                else None
            )
            rows.append(
                _make_row(
                    variable,
                    label,
                    weight,
                    0.0,
                    [],
                    {},
                    None,
                    calib_weight=calibration_weight,
                )
            )
        return rows

    def _extra_variable_rows(
        self,
        variable: str,
        weights: np.ndarray,
        total_weight: float,
        calibration_data: pl.DataFrame | None,
        calibration_weights: np.ndarray | None,
        total_calibration_weight: float,
    ) -> list[dict[str, Any]]:
        """Build summary rows for an extra configured or raw variable."""
        config = self.variable_configs.get(variable)
        is_derived = config is not None and (
            config.input_cols is not None or config.custom_transform is not None
        )
        if variable not in self.data.columns and not is_derived:
            raise ValueError(
                f"extra_vars column '{variable}' not found in data.  "
                "Register a config with input_cols or custom_transform to derive it."
            )

        if config is not None:
            preprocessor = _build_preprocessor(
                [variable], self.data, self.variable_configs
            )
            preprocessor.fit(self.data, weights=self._weights_array)
            calibration_design = (
                preprocessor.transform(calibration_data, strict=False)
                if calibration_data is not None
                else None
            )
            rows = self._encoded_variable_rows(
                variable,
                preprocessor._params.get(variable),
                config,
                preprocessor.transform(self.data),
                weights,
                total_weight,
                {},
                [],
                {},
                calibration_design,
                calibration_weights,
                total_calibration_weight,
            )
            if rows is not None:
                return rows

        if variable not in self.data.columns:
            raise ValueError(
                f"extra_vars variable '{variable}' has no binned or categorical "
                "encoding; add n_bins or encoding='onehot' to its config to "
                "resolve levels."
            )
        return self._raw_extra_variable_rows(
            variable,
            weights,
            calibration_data,
            calibration_weights,
        )

    def summary_table(
        self,
        version: str | None = None,
        calib_df: pl.DataFrame | None = None,
        extra_vars: list[str] | None = None,
    ) -> pl.DataFrame:
        """
        Summary table for all categorical and binned variables in *version*.

        Each row is one level of one discrete variable.  The dropped base
        level is included with a coefficient of zero so the full picture is
        visible at a glance.  Pure continuous variables are excluded.

        Parameters
        ----------
        version : str, optional
            Version key of the fitted model to inspect. Uses current version if none specified.
        calib_df : pl.DataFrame, optional
            An independent DataFrame (e.g. a calibration or holdout set).
            When supplied, a ``calib_weight`` column is added showing the
            total exposure weight from *calib_df* assigned to each level.
            The same ``weight_col`` used for training is read from this
            DataFrame; if absent, unit weights are assumed.
        extra_vars : list of str, optional
            Additional variables to append even if not in the model.  Each
            variable gets weight-by-level rows with ``train_coef=0.0`` and no
            fold columns.  Variables already present in the model are silently
            skipped.  If a variable has a registered config that produces a
            binned or one-hot encoding, a fresh preprocessor is fit to resolve
            proper labels; otherwise levels are resolved directly from the raw
            column (continuous variables are auto-quantile-binned into 10
            groups).  Variables that do not exist as raw columns are supported
            when their config specifies ``input_cols`` or ``custom_transform``
            (the preprocessor derives them on-the-fly), but they must produce a
            binned or categorical encoding — a plain continuous derived
            variable will raise an error.

        Returns
        -------
        pl.DataFrame
            Columns: ``variable``, ``level``, ``weight``
            [, ``calib_weight``], ``train_coef`` [, ``fold_{k}`` …].
        """
        mv = self._get_version(version or self.current_version)
        if isinstance(mv, FactorModelVersion):
            raise TypeError(
                f"Version '{version}' is an Excel factor model; "
                "summary_table is not applicable (the factor table IS the relativity table)."
            )
        preprocessor = mv.preprocessor
        weights = self._weights_array
        if weights is None:
            weights = np.ones(len(self.data))
        total_weight = float(weights.sum())
        coefficients = {
            row["feature"]: row["coefficient"]
            for row in mv.coefficients.to_dicts()
            if row["feature"] != "intercept"
        }
        fold_names, fold_coefficients = (
            self._get_fold_info(mv) if self.cv_column is not None else ([], {})
        )
        calibration = self._get_calib_arrays(preprocessor, calib_df)
        rows = self._model_summary_rows(
            mv,
            preprocessor.transform(self.data),
            weights,
            total_weight,
            coefficients,
            fold_names,
            fold_coefficients,
            *calibration,
        )
        for variable in extra_vars or []:
            if variable in set(mv.variables):
                continue
            rows.extend(
                self._extra_variable_rows(
                    variable,
                    weights,
                    total_weight,
                    calib_df,
                    calibration[1],
                    calibration[2],
                )
            )
        return pl.DataFrame(rows) if rows else pl.DataFrame()

    # ── AvE data table ───────────────────────────────────────────────────────

    def _glm_factor_arrays(
        self,
        mv: ModelVersion,
    ) -> dict[str, np.ndarray]:
        """Per-row factor array for each variable in a fitted GLM."""
        prep = mv.preprocessor
        Xt_df = prep.transform(self.data)
        coef_map: dict[str, float] = {
            r["feature"]: r["coefficient"]
            for r in mv.coefficients.to_dicts()
            if r["feature"] != "intercept"
        }
        use_exp = getattr(mv, "link", "log") == "log"
        n = len(self.data)
        result: dict[str, np.ndarray] = {}

        for var_col in mv.variables:
            if var_col not in prep.configs:
                continue
            p = prep._params.get(var_col)
            cfg = prep.configs[var_col]

            if isinstance(p, FittedCategoricalParams) and p.encoding == "onehot":
                feats = [f"{var_col}_{cat}" for cat in p.categories]
            elif isinstance(p, FittedBinnedNumericParams):
                dropped_bin = p.dropped_bin
                all_labels = p.bin_labels
                feats = []
                if f"{var_col}_missing" in Xt_df.columns:
                    feats.append(f"{var_col}_missing")
                feats += [
                    f"{var_col}_{lbl}"
                    for i, lbl in enumerate(all_labels)
                    if i != dropped_bin
                ]
            else:
                feats = [var_col] + [f"{var_col}^{d}" for d in range(2, cfg.degree + 1)]

            linear = np.zeros(n, dtype=float)
            for feat in feats:
                c = coef_map.get(feat, 0.0)
                if c != 0.0 and feat in Xt_df.columns:
                    linear += c * Xt_df[feat].to_numpy().astype(float)
            result[var_col] = np.exp(linear) if use_exp else linear

        return result

    def _factor_model_factor_arrays(
        self,
        mv: FactorModelVersion,
    ) -> dict[str, np.ndarray]:
        """Per-row factor array for each variable in a factor-table model."""
        prep = mv.preprocessor
        Xt: pl.DataFrame | None = None
        if prep is not None and mv.preprocessor_vars:
            Xt = prep.transform(self.data)

        n = len(self.data)
        factor_by_var = {
            grp: df.select(["Level", "Factor"])
            for (grp,), df in mv.factor_table.group_by("Variable")
        }
        result: dict[str, np.ndarray] = {}

        for V in mv.variables:
            ft_v = factor_by_var.get(V, pl.DataFrame({"Level": [], "Factor": []}))
            level_arr = self._resolve_factor_model_levels(V, mv, prep, Xt, n)

            tmp = pl.DataFrame({"Level": pl.Series("_l", level_arr)}).join(
                ft_v, on="Level", how="left"
            )
            result[V] = tmp["Factor"].fill_null(1.0).to_numpy().astype(float)

        return result

    def _resolve_factor_model_levels(
        self,
        V: str,
        mv: FactorModelVersion,
        prep: Any | None,
        Xt: pl.DataFrame | None,
        n: int,
    ) -> np.ndarray:
        """Resolve level strings for one variable, mirroring FactorModelVersion.predict."""
        if V in mv.preprocessor_vars and Xt is not None:
            p = prep._params[V]
            if isinstance(p, FittedCategoricalParams) and p.encoding == "onehot":
                dropped = p.dropped_category or ""
                level_arr = np.full(n, f"{dropped} (base)", dtype=object)
                for cat in p.categories:
                    feat = f"{V}_{cat}"
                    if feat in Xt.columns:
                        level_arr[Xt[feat].to_numpy().astype(bool)] = str(cat)
                return level_arr
            if isinstance(p, FittedBinnedNumericParams):
                dropped_bin = p.dropped_bin
                all_labels = p.bin_labels
                level_arr = np.full(
                    n, f"{all_labels[dropped_bin]} (base)", dtype=object
                )
                missing_feat = f"{V}_missing"
                if missing_feat in Xt.columns:
                    level_arr[Xt[missing_feat].to_numpy().astype(bool)] = "Missing"
                for i, label in enumerate(all_labels):
                    if i == dropped_bin:
                        continue
                    feat = f"{V}_{label}"
                    if feat in Xt.columns:
                        level_arr[Xt[feat].to_numpy().astype(bool)] = label
                return level_arr

        # Direct string match on raw column
        return np.array(self.data[V].cast(pl.String).to_list(), dtype=object)

    def decile_lift_chart(
        self,
        version: str | None = None,
        n_bins: int = 10,
        state_column: str | None = None,
    ) -> plt.Figure:
        """
        Decile lift chart for *version*.

        The data is sorted by predicted value and split into *n_bins* equal-sized
        groups.  The mean actual and predicted values are plotted for each group
        to show how well the model discriminates between high- and low-risk
        segments.

        Parameters
        ----------
        version : str, optional
            Model version name.  Uses current version if ``None``.
        n_bins : int
            Number of equal-sized groups to split the data into (default 10).
        state_column : str, optional
            When supplied, predictions are off-balanced within each state
            before computing the lift chart.

        Returns
        -------
        plt.Figure
            Line chart with one line each for actual and predicted.
        """
        mv = self._get_version(version or self.current_version)
        y = self._y_array
        w = self._weights_array
        p = mv.train_predictions
        states = self.data[state_column].to_numpy() if state_column else None

        fig = decile_lift_chart(y, p, w, n_bins, mv.name, states)

        return fig

    def _prediction_source(self, source: str) -> np.ndarray:
        """Resolve predictions from a registered version or numeric data column."""
        if source in self.model_versions:
            return self._get_version(source).train_predictions
        if source in self.data.columns:
            series = self.data[source]
            if not series.dtype.is_numeric():
                raise ValueError(
                    f"Column '{source}' has dtype {series.dtype}; "
                    "predictions must be numeric."
                )
            return series.cast(pl.Float64).to_numpy()
        self._get_version(source)
        return np.array([])

    def _comparison_predictions(
        self,
        source: str | None,
        state_column: str | None,
        weights: np.ndarray,
    ) -> np.ndarray | None:
        """Resolve and optionally off-balance comparison predictions."""
        resolved_source = source if source is not None else self.base_version
        if resolved_source is None:
            return None
        try:
            predictions = self._prediction_source(resolved_source)
        except KeyError as exc:
            raise ValueError(
                f"compare_version '{resolved_source}' not found in model versions "
                "or data columns."
            ) from exc
        if state_column is not None:
            return _off_balance_by_state(
                predictions,
                weights,
                self.data[state_column],
            )
        return predictions

    def _ave_levels(
        self,
        variable: str,
        synthetic_levels: dict[str, np.ndarray],
        preprocessor: Preprocessor | None,
        n_bins: int,
    ) -> np.ndarray:
        """Resolve one analysis variable into AvE level labels."""
        if variable in synthetic_levels:
            return synthetic_levels[variable]
        if variable not in self.data.columns and variable in self.variable_configs:
            config = self.variable_configs[variable]
            if config.input_cols is None and config.custom_transform is None:
                raise ValueError(
                    f"Variable '{variable}' is not a data column and has no "
                    "input_cols or custom_transform; cannot resolve levels."
                )
            variable_preprocessor = _build_preprocessor(
                [variable], self.data, self.variable_configs
            )
            variable_preprocessor.fit(self.data, weights=self._weights_array)
            return _resolve_level(
                variable,
                self.data,
                variable_preprocessor,
                n_bins,
            )
        return _resolve_level(variable, self.data, preprocessor, n_bins)

    @staticmethod
    def _aggregate_ave_levels(
        variable: str,
        levels: np.ndarray,
        data: dict[str, Any],
        aggregations: list[pl.Expr],
        column_order: list[str],
    ) -> pl.DataFrame:
        """Aggregate and naturally order one AvE table section."""
        summary = (
            pl.DataFrame({"_level": levels, **data})
            .group_by("_level")
            .agg(aggregations)
        )
        labels = _sort_labels(summary["_level"].to_list())
        order = pl.DataFrame({"_level": labels, "_order": range(len(labels))})
        return (
            summary.join(order, on="_level")
            .sort("_order")
            .drop("_order")
            .with_columns(pl.lit(variable).alias("variable"))
            .rename({"_level": "level"})
            .select(column_order)
        )

    def ave_table(
        self,
        variables: list[str],
        version: str | None = None,
        n_bins: int = 10,
        state_column: str | None = None,
        compare_version: str | None = None,
    ) -> pl.DataFrame:
        """
        Actual-vs-Expected breakdown table for a list of analysis variables.

        For each analysis variable and each of its levels, returns the total
        weighted loss (actual), weighted prediction, exposure weight, and one
        column per model variable showing ``sum(factor_i * weight_i)``.

        The factor for a model variable at row *i* is:

        - **GLM (ModelVersion)**: ``exp(linear_contribution)`` for log-link,
          where the linear contribution is the sum of ``coef * feature_value``
          across all design-matrix features belonging to that variable
          (including polynomial terms for higher-degree continuous variables).
        - **Factor model (FactorModelVersion)**: the factor looked up directly
          from the factor table for the row's level.

        Parameters
        ----------
        variables : list of str
            Analysis variables to break down by.  Need not be model variables.
        version : str, optional
            Model version name. Uses current version if none specified.
        n_bins : int
            Quantile bins for continuous non-binned analysis variables.
        state_column : str, optional
            When supplied, predictions are off-balanced within each state
            before aggregating factors.
        compare_version : str, optional
            If supplied, a second model version or column in the data whose
            scores are compared in the AvE charts.

        Returns
        -------
        pl.DataFrame
            Columns: ``variable``, ``level``, ``weight``, ``loss``,
            ``prediction``, then ``{model_var}_factor`` for each model
            variable.
        """
        model = self._get_version(version or self.current_version)
        preprocessor = getattr(model, "preprocessor", None)
        weights = (
            self._weights_array
            if self._weights_array is not None
            else np.ones(len(self.data))
        )
        predictions = model.train_predictions
        if state_column is not None:
            predictions = _off_balance_by_state(
                predictions,
                weights,
                self.data[state_column],
            )
        comparison = self._comparison_predictions(
            compare_version,
            state_column,
            weights,
        )
        factor_arrays = (
            self._factor_model_factor_arrays(model)
            if isinstance(model, FactorModelVersion)
            else self._glm_factor_arrays(model)
        )
        model_variables = [
            variable for variable in model.variables if variable in factor_arrays
        ]
        weighted_data: dict[str, Any] = {
            "_w": weights,
            "_yw": self._y_array * weights,
            "_pw": predictions * weights,
            **{
                f"_f_{variable}": factor_arrays[variable] * weights
                for variable in model_variables
            },
        }
        if comparison is not None:
            weighted_data["_cpw"] = comparison * weights

        synthetic_levels = {
            "prediction_decile": _weighted_decile_levels(predictions, weights)
        }
        if comparison is not None:
            synthetic_levels["compare_prediction_decile"] = _weighted_decile_levels(
                comparison,
                weights,
            )
        column_order = ["variable", "level", "weight", "loss", "prediction"]
        if comparison is not None:
            column_order.append("compare_prediction")
        column_order.extend(f"{variable}_factor" for variable in model_variables)
        aggregations = [
            pl.col("_w").sum().alias("weight"),
            pl.col("_yw").sum().alias("loss"),
            pl.col("_pw").sum().alias("prediction"),
            *(
                []
                if comparison is None
                else [pl.col("_cpw").sum().alias("compare_prediction")]
            ),
            *[
                pl.col(f"_f_{variable}").sum().alias(f"{variable}_factor")
                for variable in model_variables
            ],
        ]
        tables = [
            self._aggregate_ave_levels(
                variable,
                self._ave_levels(
                    variable,
                    synthetic_levels,
                    preprocessor,
                    n_bins,
                ),
                weighted_data,
                aggregations,
                column_order,
            )
            for variable in [*synthetic_levels, *variables]
        ]
        return pl.concat(tables) if tables else pl.DataFrame()

    # ── Actual vs Expected ────────────────────────────────────────────────────

    def ae_chart(
        self,
        col: str,
        version: str | None = None,
        n_bins: int = 10,
        breaks: list[float] | None = None,
        figsize: tuple[int, int] | None = None,
        state_column: str | None = None,
        show: bool = True,
    ) -> plt.Figure:
        """
        Actual vs Expected chart for *col* using model *version*.

        *col* does not need to be a model predictor.  Continuous variables
        are binned into ``n_bins`` quantile groups.  The sentinel value
        ``-999999999`` is labelled ``'Missing'``.

        Parameters
        ----------
        col : str
            Variable to slice by.
        version : str, optional
            Model version whose predictions are used.  Uses current version
            if ``None``.
        n_bins : int
            Number of quantile bins for continuous variables (default 10).
        breaks : list of float, optional
            Explicit bin edges for continuous variables.  Overrides ``n_bins``.
        figsize : tuple of int, optional
            Figure size ``(width, height)`` in inches.
        state_column : str, optional
            When supplied, predictions are off-balanced within each state
            before plotting.
        """
        mv = self._get_version(version or self.current_version)
        prep = getattr(mv, "preprocessor", None)
        if col in self.variable_configs and (prep is None or col not in prep._params):
            prep = _fit_plot_preprocessor(
                col,
                self.data,
                self.variable_configs,
                self._weights_array,
            )
        fig = ae_chart(
            X=self.data,
            y=self._y,
            col=col,
            predictions=mv.train_predictions,
            weights=self._weights,
            n_bins=n_bins,
            breaks=breaks,
            figsize=figsize,
            version_name=version,
            preprocessor=prep,
            state_column=state_column,
        )

        if show:
            plt.show()
        return fig

    def residual_chart(
        self,
        col: str,
        version: str | None = None,
        n_bins: int = 10,
        breaks: list[float] | None = None,
        figsize: tuple[int, int] | None = None,
        show: bool = True,
    ) -> plt.Figure:
        """
        Residual signal chart: ``mean_actual / mean_predicted`` per level of *col*.

        Where an A/E chart plots actual and predicted side-by-side, this chart
        shows their ratio directly.  A ratio of 1.1 / 1.05 ≈ 1.048 appears as
        a point at 1.048, making residual signal immediately readable as
        deviation from the 1.0 reference line.

        - Values **above 1.0** → model is *under-predicting* for that group.
        - Values **below 1.0** → model is *over-predicting* for that group.

        Exposure (weight) is shown as bars on the primary axis so that the
        credibility of each ratio is visible at a glance.

        Parameters
        ----------
        col : str
            Variable to slice by.  Does not need to be a model predictor.
        version : str, optional
            Version key of the model whose predictions are used. Uses current version if none specified.
        n_bins : int
            Number of quantile bins for continuous variables.
        breaks : list of float, optional
            Explicit bin edges for continuous variables.  Overrides ``n_bins``.
        figsize : tuple of int, optional
            Figure size ``(width, height)`` in inches.
        """
        mv = self._get_version(version or self.current_version)
        prep = getattr(mv, "preprocessor", None)
        if col in self.variable_configs and (prep is None or col not in prep._params):
            prep = _fit_plot_preprocessor(
                col,
                self.data,
                self.variable_configs,
                self._weights_array,
            )
        fig = residual_chart(
            X=self.data,
            y=self._y,
            col=col,
            predictions=mv.train_predictions,
            weights=self._weights,
            n_bins=n_bins,
            breaks=breaks,
            figsize=figsize,
            version_name=version,
            preprocessor=prep,
        )

        if show:
            plt.show()
        return fig

    def plot_all_variables(
        self,
        version: str | None = None,
        chart: str = "residual",
        n_bins: int = 10,
        figsize: tuple[int, int] | None = None,
        show: bool = True,
    ) -> list[plt.Figure]:
        """
        Plot a residual or A/E chart for every variable in *version*.

        Parameters
        ----------
        version : str, optional
            Version key whose variable list drives the loop. Uses current version if none specified.
        chart : {'residual', 'ae'}
            ``'residual'`` (default) — ``mean_actual / mean_predicted`` per
            level, with a horizontal reference line at 1.0.
            ``'ae'`` — side-by-side actual vs expected bars.
        n_bins : int
            Number of quantile bins used for continuous variables.
        figsize : tuple of int, optional
            Figure size ``(width, height)`` in inches applied to each chart.
        show : bool
            Call ``plt.show()`` after each chart.

        Returns
        -------
        list of matplotlib.figure.Figure
            One figure per variable, in the same order as ``mv.variables``.
        """
        if chart not in ("residual", "ae"):
            raise ValueError(f"chart must be 'residual' or 'ae', got {chart!r}")
        version = version or self.current_version
        mv = self._get_version(version)
        figs: list[plt.Figure] = []
        for col in mv.variables:
            if chart == "ae":
                fig = self.ae_chart(
                    col, version=version, n_bins=n_bins, figsize=figsize, show=False
                )
            else:
                fig = self.residual_chart(
                    col, version=version, n_bins=n_bins, figsize=figsize, show=False
                )
            figs.append(fig)
            if show:
                plt.show()
        return figs

    # ── Model comparison ──────────────────────────────────────────────────────

    def _get_dl_score(self, y, p1, p2, weights, n_buckets, deviation="absolute"):
        dl_data = double_lift_table(y, p1, p2, weights=weights, n_buckets=n_buckets)
        return dl_data, double_lift_score(dl_data, deviation=deviation)

    def compare_models(
        self,
        version1: str | None = None,
        version2: str | None = None,
        n_buckets: int = 10,
        figsize: tuple[int, int] | None = None,
        dl_deviation: str | None = "absolute",
        state_column: str | None = None,
        show: bool = True,
    ) -> dict[str, Any]:
        """
        Compare two model versions: metrics table and double-lift chart.

        ``version2`` can be either:

        * A fitted model version name registered with :meth:`fit_model` or
          :meth:`add_excel_version` (the original behaviour), **or**
        * A **column name** in the tool's dataset whose values are pre-computed
          predictions (e.g. an incumbent / external model stored in the data).
        * A preset version using the set_base_version method, which also registers the version name.

        Registered model versions take priority: if a name matches both a
        version and a column, the version is used.

        Parameters
        ----------
        version1 : str, optional
            First model version (must be a registered version name). Uses current version if none specified.
        version2 : str, optional
            Second model — either a registered version name or a dataframe
            column containing predictions.
        n_buckets : int
            Number of buckets for the double-lift table.
        figsize : tuple of int, optional
            Figure size ``(width, height)`` in inches for the double-lift chart.
        dl_deviation : {'absolute', 'relative'}, optional
            Deviation metric used for the double-lift score.
        state_column : str, optional
            When supplied, predictions are off-balanced within each state
            before comparison.
        show : bool
            Call ``plt.show()`` after rendering the chart.

        Returns
        -------
        dict
            ``{'metrics': pl.DataFrame, 'double_lift': pl.DataFrame}``
        """
        mv1 = self._get_version(version1 or self.current_version)

        y = self._y_array
        w = self._weights_array
        p1 = mv1.train_predictions

        # Resolve version2: registered model version takes priority over column.
        if version2 is None and self.base_version is None:
            raise ValueError("version2 must be specified if no base version is set.")
        version2 = version2 or self.base_version

        p2 = self._prediction_source(version2)

        if state_column is not None:
            states = self.data[state_column]
            p1 = _off_balance_by_state(p1, w, states)
            p2 = _off_balance_by_state(p2, w, states)

        dl_data, dl_sc = self._get_dl_score(
            y, p1, p2, weights=w, n_buckets=n_buckets, deviation=dl_deviation
        )

        metrics = compare_metrics(
            y,
            p1,
            p2,
            weights=w,
            name1=version1,
            name2=version2,
            dl_score=dl_sc,
        )

        print("\n" + "=" * 60)
        print(f"  Comparison: {version1}  vs  {version2}")
        print("=" * 60)
        print(metrics)
        # print("=" * 60)
        # if dl_sc < 0:
        #     dl_interp = f"negative -> {version1} wins"
        # elif dl_sc > 0:
        #     dl_interp = f"positive -> {version2} wins"
        # else:
        #     dl_interp = "tie"
        # print(f"  double_lift_score interpretation: {dl_interp}")
        # print("=" * 60 + "\n")

        double_lift_chart(
            y,
            p1,
            p2,
            weights=w,
            n_buckets=n_buckets,
            name1=version1,
            name2=version2,
            figsize=figsize,
        )

        if show:
            plt.show()

        return {"metrics": metrics, "double_lift": dl_data}

    def midpoint_movement(
        self,
        expense_values: dict[str, float],
        rate_adequacy_factors: pl.DataFrame,
        state_column: str = "PREDOM_STATE_ABBREV",
        version1: str | None = None,
        version2: str | None = None,
        extra_cols: list[str] | None = None,
        show_plot: bool = True,
    ) -> pl.DataFrame:
        """
        Compute the midpoint movement between two model versions. Assumes moving from version1 to version2.

        Parameters
        ----------
        expense_values : dict
            Dictionary of expense values. Need CWR_FIXED_EXP_RATIO, CWR_ULAE_RATIO, and CWR_LOSS_PLUS_ALAE_RATIO
        rate_adequacy_factors : pl.DataFrame
            DataFrame containing rate adequacy factors for each state. State column should be named 'STATE' (two letter code)
            and rate adequacy factor column should be named 'RATE_ADEQUACY_FACTOR'.
        state_column : str
            Name of the column in the data that contains state information. Default is 'PREDOM_STATE_ABBREV'
        version1 : str, optional
            First model version (must be a registered version name). Uses current version if none specified.
        version2 : str, optional
            Second model — either a registered version name or a dataframe
            column containing predictions.
        extra_cols : list of str, optional
            Additional columns to include in the output DataFrame. These columns must exist in the data.
        show_plot : bool
            Display a histogram of midpoint movements after computing (default ``True``).

        Returns
        -------
        pl.DataFrame
            DataFrame with columns: ``version1``, ``version2``, ``midpoint_movement``.
        """
        if not all(
            k in expense_values
            for k in [
                "CWR_FIXED_EXP_RATIO",
                "CWR_ULAE_RATIO",
                "CWR_LOSS_PLUS_ALAE_RATIO",
            ]
        ):
            raise ValueError(
                "expense_values must contain CWR_FIXED_EXP_RATIO, CWR_ULAE_RATIO, and CWR_LOSS_PLUS_ALAE_RATIO"
            )

        if (
            "STATE" not in rate_adequacy_factors.columns
            or "RATE_ADEQUACY_FACTOR" not in rate_adequacy_factors.columns
        ):
            raise ValueError(
                "rate_adequacy_factors must contain 'STATE' and 'RATE_ADEQUACY_FACTOR' columns"
            )

        mv1 = self._get_version(version1 or self.current_version)
        p1 = mv1.train_predictions

        # Resolve version2: registered model version takes priority over column.
        if version2 is None and self.base_version is None:
            raise ValueError("version2 must be specified if no base version is set.")
        version2 = version2 or self.base_version

        p2 = self._prediction_source(version2)

        w = self._weights_array

        policy_cols = ["POLICY_NUM", "POLICY_EFF_DT", state_column]
        if any(col not in self.data.columns for col in policy_cols):
            raise ValueError(f"Data must contain columns: {policy_cols}")

        if extra_cols is not None:
            missing_cols = [col for col in extra_cols if col not in self.data.columns]
            if missing_cols:
                raise ValueError(f"Extra columns {missing_cols} not found in the data.")
            policy_cols += extra_cols

        policy_info = self.data.select(policy_cols)

        midpoint_movement_df = compute_midpoint_movement(
            policy_info,
            p1,
            p2,
            expense_values,
            rate_adequacy_factors,
            weights=w,
            name1=mv1.name,
            name2=version2,
        )

        if show_plot:
            midpoint_movement_histogram(
                midpoint_movement_df["midpoint_movement"].to_numpy(), w
            )

        return midpoint_movement_df

    def list_versions(self) -> pl.DataFrame:
        """Summary table of all stored model versions."""
        rows = []
        y = self._y_array
        w = self._weights_array
        for name, mv in self.model_versions.items():
            m = compute_metrics(y, mv.train_predictions, w, name)
            metric_vals = {
                r: v for r, v in zip(m["metric"].to_list(), m[name].to_list())
            }
            rows.append(
                {
                    "version": name,
                    "n_variables": len(mv.variables),
                    "alpha": mv.alpha,
                    "l1_ratio": mv.l1_ratio,
                    "n_nonzero": max(
                        0, int((mv.coefficients["coefficient"] != 0).sum()) - 1
                    ),
                    "rmse": metric_vals.get("rmse", float("nan")),
                    "mae": metric_vals.get("mae", float("nan")),
                    "gini_norm": metric_vals.get("gini_norm", float("nan")),
                }
            )
        return pl.DataFrame(rows) if rows else pl.DataFrame()

    # ── Persistence ───────────────────────────────────────────────────────────

    def save(
        self,
        version: str | None = None,
        filepath: str | Path | None = "models/model.pkl",
    ) -> None:
        """
        Save a model version to *filepath* (pickle).

        Parameters
        ----------
        version : str, optional
            Version key to save.  Uses current version if ``None``.
        filepath : str
            Destination path (e.g. ``'models/v1.pkl'``).
        """
        mv = self._get_version(version or self.current_version)
        save_version(mv, self, filepath)

    @classmethod
    def load_version(
        cls,
        filepath: str | Path,
        data: pl.DataFrame | None = None,
        target_col: str | None = None,
        weight_col: str | None = None,
        cv_column: str | None = None,
        version_name: str | None = None,
        into: ModelingTool | None = None,
    ) -> ModelingTool:
        """
        Load and refit a saved version, creating or updating a tool.

        When *into* is omitted, *data* is required and a new tool is returned.
        When *into* is supplied, its data and settings are used, the loaded
        version is appended atomically, and the same tool is returned.

        Parameters
        ----------
        filepath : str
            Path to the saved ``.pkl`` file.
        data : pl.DataFrame
            Training data to refit on (must contain the same columns).
        target_col : str, optional
            Override the saved target column name.
        weight_col : str, optional
            Override the saved weight column name.
        cv_column : str, optional
            Column whose unique values define predefined CV folds.
        version_name : str, optional
            Registration name. Uses the saved name when omitted.
        into : ModelingTool, optional
            Existing compatible tool to receive the loaded version.
        """
        if into is not None and not isinstance(into, cls):
            raise TypeError("into must be a ModelingTool or None.")
        if into is not None and any(
            value is not None for value in (data, target_col, weight_col, cv_column)
        ):
            raise ValueError(
                "data, target_col, weight_col, and cv_column cannot be supplied "
                "when into is provided; the destination tool's settings are used."
            )
        if into is None and not isinstance(data, pl.DataFrame):
            raise TypeError(
                "data must be a polars DataFrame when into is not provided."
            )

        snapshot = _load_version_snapshot(filepath, data=None, refit=False)["snapshot"]
        version = snapshot["version"]
        settings = snapshot["tool_settings"]
        loaded_name = version_name or version["name"]
        if snapshot.get("version_type") == "factor":
            raise ValueError(
                "Cannot refit a factor model version — factor tables are not refittable from data. "
                "Use ModelingTool.load_version_frozen() to restore this version."
            )

        if into is None:
            tool = cls(
                data=data,
                target_col=target_col or settings["target_col"],
                weight_col=weight_col or settings["weight_col"],
                offset_col=settings.get("offset_col"),
                link=settings["link"],
                drop_reference=settings.get("drop_reference", "max_weight"),
                cv_column=cv_column,
            )
            tool.variable_configs.update(settings["variable_configs"])
            _fit_loaded_snapshot(tool, snapshot, loaded_name)
        else:
            _validate_load_destination(into, settings, loaded_name)
            merged_configs = _merge_variable_configs(into, settings)
            staged = cls(
                data=into.data,
                target_col=into.target_col,
                weight_col=into.weight_col,
                offset_col=into.offset_col,
                link=into.link,
                tweedie_power=into.tweedie_power,
                drop_reference=into.drop_reference,
                cv_column=into.cv_column,
            )
            staged.variable_configs.update(merged_configs)
            _fit_loaded_snapshot(staged, snapshot, loaded_name)
            for col, config in settings["variable_configs"].items():
                into.variable_configs.setdefault(col, config)
            into.model_versions[loaded_name] = staged.model_versions[loaded_name]
            into.current_version = loaded_name
            tool = into

        print(
            f"Loaded '{version['name']}' from {filepath!r}, refitted as version '{loaded_name}'."
        )
        return tool

    @classmethod
    def load_version_frozen(
        cls,
        filepath: str | Path,
        data: pl.DataFrame | None = None,
        version_name: str | None = None,
        into: ModelingTool | None = None,
    ) -> ModelingTool:
        """
        Restore a fitted version, creating or updating a tool without refitting.

        When *into* is supplied, its data is used for training predictions and
        the reconstructed version is appended atomically.

        Parameters
        ----------
        filepath : str
            Path to the saved ``.pkl`` file.
        data : pl.DataFrame, optional
            Dataset used to compute ``train_predictions`` for the loaded
            version. When omitted, the model is restored with empty training
            predictions and can still score later via ``predict(data)``.
        version_name : str, optional
            Registration name. Uses the saved name when omitted.
        into : ModelingTool, optional
            Existing compatible tool to receive the loaded version.
        """
        if into is not None and not isinstance(into, cls):
            raise TypeError("into must be a ModelingTool or None.")
        if into is not None and data is not None:
            raise ValueError(
                "data cannot be supplied when into is provided; "
                "the destination tool's data is used."
            )
        if data is not None and not isinstance(data, pl.DataFrame):
            raise TypeError("data must be a polars DataFrame or None.")

        snapshot = _load_version_snapshot(filepath, data=None, refit=False)["snapshot"]
        version = snapshot["version"]
        settings = snapshot["tool_settings"]
        loaded_name = version_name or version["name"]
        if into is None:
            tool = _new_frozen_tool(cls, snapshot, data)
            model = _restore_frozen_model(
                snapshot, loaded_name, tool.data, compute_predictions=data is not None
            )
        else:
            _validate_load_destination(into, settings, loaded_name)
            tool = into
            model = _restore_frozen_model(
                snapshot,
                loaded_name,
                tool.data,
                compute_predictions=bool(tool.data.columns),
            )
        for col, config in settings["variable_configs"].items():
            tool.variable_configs.setdefault(col, config)
        tool.model_versions[loaded_name] = model
        tool.current_version = loaded_name
        if snapshot.get("version_type") != "factor":
            tool.family = version["family"]
        print(
            f"Loaded frozen '{version['name']}' from {filepath!r} as '{loaded_name}'."
        )
        return tool

    @classmethod
    def load(
        cls,
        filepath: str | Path,
        data: pl.DataFrame,
        target_col: str | None = None,
        weight_col: str | None = None,
        cv_column: str | None = None,
        version_name: str | None = None,
    ) -> ModelingTool:
        """Backward-compatible alias for :meth:`load_version`."""
        return cls.load_version(
            filepath,
            data=data,
            target_col=target_col,
            weight_col=weight_col,
            cv_column=cv_column,
            version_name=version_name,
        )

    @classmethod
    def load_frozen(
        cls,
        filepath: str | Path,
        data: pl.DataFrame | None = None,
    ) -> ModelingTool:
        """Backward-compatible frozen loader that registers as ``'v1'``."""
        return cls.load_version_frozen(filepath, data=data, version_name="v1")

    @classmethod
    def load_from_excel(
        cls,
        excel_path: str | Path,
        sheet_name: str,
        data: pl.DataFrame,
        target_col: str,
        weight_col: str | None = None,
        pkl_path: str | None = None,
        version: str = "excel",
        missing_factor: float = 1.0,
        offset_col: str | None = None,
    ) -> ModelingTool:
        """
        Build a :class:`ModelingTool` from an Excel factor table.

        If *pkl_path* is supplied the saved model is loaded frozen (providing
        preprocessing context for numeric/binned variables).  Without a pkl the
        tool works in standalone mode — only categorical or pre-banded string
        columns are resolved directly.

        Parameters
        ----------
        excel_path : str
            Path to the ``.xlsx`` workbook.
        sheet_name : str
            Sheet containing ``Variable``, ``Level``, ``Factor`` columns.
        data : pl.DataFrame
            Dataset to score.
        target_col : str
            Name of the target column (required to construct the tool).
        weight_col : str, optional
            Name of the exposure-weight column.
        pkl_path : str, optional
            Path to a saved ``ModelingTool`` pickle (from :meth:`save`).
            When provided, the frozen model's preprocessors are used for
            level resolution of binned/categorical variables.
        version : str
            Version label for the Excel model (default ``'excel'``).
        missing_factor : float
            Factor applied to unseen levels (default 1.0, with a warning).
        offset_col : str, optional
            Column in *data* containing a per-row multiplicative offset on
            the **response scale** (i.e. already exponentiated) for the Excel
            factor-table version.

        Returns
        -------
        ModelingTool
            Contains *data* and a single registered version *version*.
        """
        if not isinstance(data, pl.DataFrame):
            raise TypeError("data must be a polars DataFrame.")

        if pkl_path is not None:
            # Load the saved model frozen to get preprocessor + variable configs
            frozen = cls.load_frozen(pkl_path, data)
            tool = cls(
                data=data,
                target_col=target_col or frozen.target_col,
                weight_col=weight_col or frozen.weight_col,
            )
            tool.variable_configs = frozen.variable_configs
            tool.model_versions = frozen.model_versions  # 'v1' has the preprocessor
        else:
            tool = cls(data=data, target_col=target_col, weight_col=weight_col)

        tool.add_excel_version(
            excel_path,
            sheet_name,
            version=version,
            missing_factor=missing_factor,
            offset_col=offset_col,
        )
        return tool

    # ── Discovery ─────────────────────────────────────────────────────────────

    def fit_shadow_gbm(
        self,
        feature_cols: list[str] | None = None,
        tweedie_power: float | None = 1.50,
        **kwargs: Any,
    ) -> Any:
        """
        Fit a LightGBM on raw features for diagnostic purposes.

        Stores the fitted model on ``self._shadow_model``.  Required before
        calling :meth:`interaction_ranking`, :meth:`shap_importance`, and
        other shadow-GBM methods.

        Parameters
        ----------
        feature_cols : list of str, optional
            Columns to use as features.  Defaults to all numeric columns
            (excluding target, weight, and offset).
        tweedie_power : float, optional
            Tweedie variance power for the GBM objective (default 1.5).
        **kwargs
            Additional keyword arguments forwarded to LightGBM.
        """
        from .discovery import fit_shadow_gbm

        model = fit_shadow_gbm(
            self.data,
            self.target_col,
            weight_col=self.weight_col,
            offset_col=self.offset_col,
            feature_cols=feature_cols,
            tweedie_power=tweedie_power,
            variable_configs=self.variable_configs,
            **kwargs,
        )
        self._shadow_model = model
        print(f"  Shadow GBM fitted on {len(model._shadow_feature_cols)} features.")
        return model

    def interaction_ranking(self, top_n: int = 20, **kwargs: Any) -> pl.DataFrame:
        """Rank variable pairs by H-statistic.  Requires :meth:`fit_shadow_gbm` first.

        Parameters
        ----------
        top_n : int
            Number of top variable pairs to return (default 20).
        **kwargs
            Additional arguments forwarded to the underlying ranking function.
        """
        from .discovery import interaction_ranking

        if not hasattr(self, "_shadow_model"):
            raise RuntimeError("Call fit_shadow_gbm() first.")
        return interaction_ranking(
            self._shadow_model,
            self.data,
            top_n=top_n,
            **kwargs,
        )

    def partial_dependence_2d(
        self, var1: str, var2: str, **kwargs: Any
    ) -> pl.DataFrame:
        """2D partial dependence for a variable pair.  Requires :meth:`fit_shadow_gbm` first."""
        from .discovery import partial_dependence_2d

        if not hasattr(self, "_shadow_model"):
            raise RuntimeError("Call fit_shadow_gbm() first.")
        return partial_dependence_2d(
            self._shadow_model, self.data, var1, var2, **kwargs
        )

    def permutation_importance(
        self,
        version: str | None = None,
        metric_fn: Any | None = None,
        **kwargs: Any,
    ) -> pl.DataFrame:
        """
        Permutation importance for the shadow GBM.

        Requires :meth:`fit_shadow_gbm` to be called first.

        Parameters
        ----------
        version : str, optional
            Not currently supported.  Must be ``None``.
        metric_fn : callable, optional
            Custom metric function ``(y_true, y_pred, weights) -> float``.
        **kwargs
            Additional arguments forwarded to the underlying function.
        """
        from .discovery import permutation_importance as _perm_imp

        if version is not None:
            raise NotImplementedError(
                "Permutation importance on fitted GLMs requires the full "
                "transform pipeline.  Use with shadow GBM instead, or pass "
                "a version=None to use the shadow model."
            )
        else:
            if not hasattr(self, "_shadow_model"):
                raise RuntimeError("Call fit_shadow_gbm() first.")
            return _perm_imp(
                self._shadow_model,
                self.data,
                self.target_col,
                weight_col=self.weight_col,
                metric_fn=metric_fn,
                **kwargs,
            )

    def shap_importance(
        self,
        feature_cols: list[str] | None = None,
        sample_size: int = 500,
        random_state: int = 42,
    ) -> pl.DataFrame:
        """
        SHAP-based feature importance using TreeExplainer.

        Requires :meth:`fit_shadow_gbm` first and ``pip install shap``.

        Parameters
        ----------
        feature_cols : list of str, optional
            Columns to explain.  Defaults to those used by the shadow GBM.
        sample_size : int
            Number of rows sampled for SHAP computation (default 500).
        random_state : int
            Random seed for sampling (default 42).

        Returns
        -------
        pl.DataFrame
            Columns: ``variable``, ``importance_mean``, ``importance_std``.
        """
        from .discovery import shap_importance as _shap_importance

        if not hasattr(self, "_shadow_model"):
            raise RuntimeError("Call fit_shadow_gbm() first.")
        return _shap_importance(
            self._shadow_model,
            self.data,
            feature_cols=feature_cols,
            sample_size=sample_size,
            random_state=random_state,
        )

    def shap_dependence(
        self,
        var: str,
        color_var: str | None = None,
        feature_cols: list[str] | None = None,
        sample_size: int = 500,
        random_state: int = 42,
    ) -> pl.DataFrame:
        """
        SHAP dependence data for *var* — reveals transform shape and breakpoints.

        Requires :meth:`fit_shadow_gbm` first and ``pip install shap``.

        Parameters
        ----------
        var : str
            Primary feature to plot.
        color_var : str, optional
            Secondary feature used to colour the scatter points, revealing
            interaction effects.
        feature_cols : list of str, optional
            Columns to explain.  Defaults to those used by the shadow GBM.
        sample_size : int
            Number of rows sampled for SHAP computation (default 500).
        random_state : int
            Random seed for sampling (default 42).

        Returns
        -------
        pl.DataFrame
            Columns: ``{var}``, ``shap_value`` [, ``{color_var}``].
        """
        from .discovery import shap_dependence as _shap_dependence

        if not hasattr(self, "_shadow_model"):
            raise RuntimeError("Call fit_shadow_gbm() first.")
        return _shap_dependence(
            self._shadow_model,
            self.data,
            var,
            color_var=color_var,
            feature_cols=feature_cols,
            sample_size=sample_size,
            random_state=random_state,
        )

    def shap_interaction_ranking(
        self,
        feature_cols: list[str] | None = None,
        sample_size: int = 200,
        random_state: int = 42,
        top_n: int = 20,
    ) -> pl.DataFrame:
        """
        Rank variable pairs by SHAP interaction strength.

        Faster and more accurate than Friedman H-statistic for tree models.
        Requires :meth:`fit_shadow_gbm` first and ``pip install shap``.

        Parameters
        ----------
        feature_cols : list of str, optional
            Columns to explain.  Defaults to those used by the shadow GBM.
        sample_size : int
            Number of rows sampled for SHAP interaction values (default 200).
        random_state : int
            Random seed for sampling (default 42).
        top_n : int
            Number of top variable pairs to return (default 20).

        Returns
        -------
        pl.DataFrame
            Columns: ``var1``, ``var2``, ``interaction_strength``.
        """
        from .discovery import shap_interaction_ranking as _shap_ir

        if not hasattr(self, "_shadow_model"):
            raise RuntimeError("Call fit_shadow_gbm() first.")
        return _shap_ir(
            self._shadow_model,
            self.data,
            feature_cols=feature_cols,
            sample_size=sample_size,
            random_state=random_state,
            top_n=top_n,
        )

    def tree_interaction_cooccurrence(self, top_n: int = 20) -> pl.DataFrame:
        """
        Fast interaction ranking by tree co-occurrence weighted by split gain.

        Use as a cheap pre-screen before running SHAP interaction ranking.
        Requires :meth:`fit_shadow_gbm` first.

        Parameters
        ----------
        top_n : int
            Number of top variable pairs to return (default 20).

        Returns
        -------
        pl.DataFrame
            Columns: ``var1``, ``var2``, ``cooccurrence_score``.
        """
        from .discovery import tree_interaction_cooccurrence as _tic

        if not hasattr(self, "_shadow_model"):
            raise RuntimeError("Call fit_shadow_gbm() first.")
        return _tic(self._shadow_model, top_n=top_n)

    def suggest_category_groups(
        self,
        col: str,
        max_groups: int = 10,
        min_exposure_pct: float = 0.01,
        verbose: bool = True,
    ):
        """
        Suggest groupings for a high-cardinality categorical variable.

        Levels are sorted by exposure-weighted mean target and merged greedily
        until at most ``max_groups`` groups remain.

        Parameters
        ----------
        col : str
            Categorical column to group.
        max_groups : int
            Maximum number of groups to produce (default 10).
        min_exposure_pct : float
            Minimum share of total exposure a level must have to avoid
            automatic merging (default 0.01).
        verbose : bool
            Print the suggested mapping table (default ``True``).

        Returns
        -------
        tuple[dict[str, str], pl.DataFrame]
            ``(level_to_group, summary)`` — mapping dict and a summary table
            with columns ``group``, ``levels``, ``exposure``, ``mean_target``.
        """
        from .discovery import suggest_category_groups as _scg

        return _scg(
            col,
            self.data,
            self._y,
            weights=self._weights,
            max_groups=max_groups,
            min_exposure_pct=min_exposure_pct,
            verbose=verbose,
        )

    def monotonicity_test(
        self,
        var: str,
        feature_cols: list[str] | None = None,
        n_estimators: int = 100,
        random_state: int = 42,
        verbose: bool = True,
        **kwargs: Any,
    ) -> dict[str, Any]:
        """
        Measure the RMSE cost of enforcing a monotone constraint on *var*.

        Fits constrained GBMs (increasing and decreasing) and reports how much
        accuracy is lost versus an unconstrained baseline.  A small cost
        (< ~1 %) means the monotone constraint is safe to apply.

        Parameters
        ----------
        var : str
            Column to test for monotonicity.
        feature_cols : list of str, optional
            Feature columns for the GBM.  Defaults to all numeric columns.
        n_estimators : int
            Number of GBM trees (default 100).
        random_state : int
            Random seed (default 42).
        verbose : bool
            Print a summary of the monotonicity cost (default ``True``).
        **kwargs
            Additional arguments forwarded to LightGBM.

        Returns
        -------
        dict
            Keys: ``unconstrained_rmse``, ``constrained_rmse_pos``,
            ``constrained_rmse_neg``, ``cost_pos``, ``cost_neg``,
            ``recommended``.
        """
        from .discovery import monotonicity_test as _mt

        return _mt(
            self.data,
            self.target_col,
            var,
            weight_col=self.weight_col,
            feature_cols=feature_cols,
            n_estimators=n_estimators,
            random_state=random_state,
            variable_configs=self.variable_configs,
            verbose=verbose,
            **kwargs,
        )

    def boruta_select(
        self,
        feature_cols: list[str] | None = None,
        n_estimators: int = 100,
        n_iterations: int = 20,
        threshold: float = 0.05,
        random_state: int = 42,
        **kwargs: Any,
    ) -> pl.DataFrame:
        """
        Boruta-style feature selection using shadow (shuffled) features.

        Each real feature must beat the maximum shadow-feature importance in
        at least ``1 - threshold`` of iterations to be selected.

        Parameters
        ----------
        feature_cols : list of str, optional
            Columns to evaluate.  Defaults to all numeric columns.
        n_estimators : int
            Number of GBM trees per iteration (default 100).
        n_iterations : int
            Number of Boruta iterations (default 20).
        threshold : float
            Significance threshold; a feature is selected when its pass rate
            exceeds ``1 - threshold`` (default 0.05).
        random_state : int
            Random seed (default 42).
        **kwargs
            Additional arguments forwarded to LightGBM.

        Returns
        -------
        pl.DataFrame
            Columns: ``variable``, ``pass_rate``, ``selected``.
            Sorted by ``pass_rate`` descending.
        """
        from .discovery import boruta_select as _boruta

        return _boruta(
            self.data,
            self.target_col,
            weight_col=self.weight_col,
            feature_cols=feature_cols,
            n_estimators=n_estimators,
            n_iterations=n_iterations,
            threshold=threshold,
            random_state=random_state,
            variable_configs=self.variable_configs,
            **kwargs,
        )

    def residual_gbm(
        self,
        version: str | None = None,
        feature_cols: list[str] | None = None,
        top_n: int = 10,
        **kwargs: Any,
    ) -> pl.DataFrame:
        """
        Fit a GBM on GLM residuals to find missing signal.

        Parameters
        ----------
        version : str, optional
            Model version whose residuals to analyse.  Uses current version
            if ``None``.
        feature_cols : list of str, optional
            Raw feature columns.  Defaults to all numeric columns.
        top_n : int
            Number of top features to return (default 10).
        **kwargs
            Additional arguments forwarded to LightGBM.
        """
        from .discovery import residual_gbm as _residual_gbm

        mv = self._get_version(version or self.current_version)
        actual = self._y_array
        predicted = mv.train_predictions
        safe_pred = np.where(np.abs(predicted) < 1e-12, 1e-12, predicted)
        residuals = actual / safe_pred

        if feature_cols is None:
            exclude = {self.target_col}
            if self.weight_col:
                exclude.add(self.weight_col)
            feature_cols = [
                c
                for c in self.data.columns
                if c not in exclude and self.data[c].dtype.is_numeric()
            ]

        return _residual_gbm(
            self.data,
            residuals,
            feature_cols,
            weight_col=self.weight_col,
            offset_col=self.offset_col,
            top_n=top_n,
            variable_configs=self.variable_configs,
            **kwargs,
        )

    # ── Enhanced residual analysis ───────────────────────────────────────────

    def residual_heatmap(
        self,
        col1: str,
        col2: str,
        version: str | None = None,
        n_bins: int = 8,
        **kwargs: Any,
    ) -> tuple[plt.Figure, pl.DataFrame]:
        """
        2D residual heatmap: A/E ratio across two variable dimensions.

        Parameters
        ----------
        col1 : str
            First dimension variable.
        col2 : str
            Second dimension variable.
        version : str, optional
            Model version whose predictions are used.  Uses current version
            if ``None``.
        n_bins : int
            Number of quantile bins for continuous variables (default 8).
        **kwargs
            Additional keyword arguments forwarded to the underlying plot
            function.
        """
        from .plots import residual_heatmap as _residual_heatmap

        mv = self._get_version(version or self.current_version)
        prep = getattr(mv, "preprocessor", None)
        analysis_cols = [col1, col2]
        if any(
            col in self.variable_configs and (prep is None or col not in prep._params)
            for col in analysis_cols
        ):
            configured_cols = [
                col for col in analysis_cols if col in self.variable_configs
            ]
            prep = _build_preprocessor(
                configured_cols, self.data, self.variable_configs
            )
            prep.fit(self.data, weights=self._weights_array)
        fig, data = _residual_heatmap(
            self.data,
            self._y,
            col1,
            col2,
            predictions=mv.train_predictions,
            weights=self._weights,
            preprocessor=prep,
            n_bins=n_bins,
            **kwargs,
        )

        return fig, data

    # ── Diagnostics ──────────────────────────────────────────────────────────

    def regularization_path(
        self,
        variables: list[str] | None = None,
        version: str | None = None,
        tweedie_power: float | None = 1.50,
        l1_ratio: float = 0.5,
        n_alphas: int = 50,
        alpha_min: float = 1e-5,
        alpha_max: float = 10.0,
        show: bool = True,
    ) -> pl.DataFrame:
        """
        Fit the GLM at a sequence of alpha values and track coefficient evolution.

        Parameters
        ----------
        variables : list of str, optional
            Variables to use.  Defaults to the variables in *version*.
        version : str, optional
            Existing version to derive the variable list from.
        tweedie_power : float, optional
            Tweedie variance power used when fitting along the path (default
            1.5, or taken from *version* when supplied).
        l1_ratio : float
            Elastic-net mixing parameter held fixed across the path (default
            0.5).
        n_alphas : int
            Number of alpha values to evaluate (default 50).
        alpha_min : float
            Smallest alpha to evaluate (default 1e-5).
        alpha_max : float
            Largest alpha to evaluate (default 10.0).
        show : bool
            Display the regularization path chart (default ``True``).
        """
        from .plots import regularization_path_plot

        mv = None
        if variables is None:
            if version is not None:
                mv = self._get_version(version)
                variables = mv.variables
            else:
                raise ValueError("Provide variables or version.")

        alphas = np.logspace(np.log10(alpha_min), np.log10(alpha_max), n_alphas)[::-1]
        tweedie_power = mv.tweedie_power if mv is not None else tweedie_power
        family = mv.family if mv is not None else "tweedie"
        prep = mv.preprocessor if mv is not None else None

        rows = []
        for alpha_val in alphas:
            mv = fit_model(
                X=self.data,
                y=self._y,
                variables=variables,
                version_name="_regpath",
                configs=self.variable_configs,
                preprocessor=prep,
                weights=self._weights_array,
                offset=self._offset_array,
                family=family,
                link=self.link,
                tweedie_power=tweedie_power,
                alpha=float(alpha_val),
                l1_ratio=l1_ratio,
                use_cv=False,
                drop_reference=self.drop_reference,
            )
            coefs = mv.coefficients
            for feat_row in coefs.iter_rows(named=True):
                if feat_row["feature"] != "intercept":
                    rows.append(
                        {
                            "alpha": float(alpha_val),
                            "variable": feat_row["feature"],
                            "coefficient": feat_row["coefficient"],
                        }
                    )

        path_df = pl.DataFrame(rows)

        if show:
            _fig = regularization_path_plot(path_df)
            plt.show()

        return path_df

    def _apply_cv_fold(self, mask: pl.Series):
        """
        Shortcut for filtering the data, target, and weights to the train/test fold in cross-validation methods.
        """
        x, y = self.data.filter(mask), self._y_array[mask]
        w = self._weights_array[mask] if self._weights_array is not None else None
        o = self._offset_array[mask] if self._offset_array is not None else None
        return x, y, w, o

    def _get_cv_metrics_gini(self, train_mask, test_mask, model_version):
        """
        Fit model on the given train fold and return the gini coefficient of the test fold.  Used for cross-validation metrics in gini mode.
        """
        X_train, y_train, w_train, o_train = self._apply_cv_fold(train_mask)
        X_test, y_test, w_test, o_test = self._apply_cv_fold(test_mask)

        fold_mv = fit_model(
            X=X_train,
            y=y_train,
            variables=model_version.variables,
            version_name="_cv_fold",
            weights=w_train,
            offset=o_train,
            family=model_version.family,
            link=self.link,
            preprocessor=model_version.preprocessor,
            alpha=model_version.alpha,
            l1_ratio=model_version.l1_ratio,
            use_cv=False,
            drop_reference=self.drop_reference,
        )

        pred_test = fold_mv.predict(X_test, offset=o_test)

        return gini_coefficient(y_test, pred_test, w_test), w_test.sum()

    def _get_cv_metrics_dls(
        self,
        train_mask,
        test_mask,
        dl_base_preds,
        dl_refit_base,
        dl_base_version,
        model_version,
        dl_deviation="absolute",
    ):
        """
        Fit model on the given train fold and return the double lift score of the test fold.  Used for cross-validation metrics in double lift mode.
        """
        X_train, y_train, w_train, o_train = self._apply_cv_fold(train_mask)
        X_test, y_test, w_test, o_test = self._apply_cv_fold(test_mask)

        fold_mv = fit_model(
            X=X_train,
            y=y_train,
            variables=model_version.variables,
            version_name="_cv_fold",
            weights=w_train,
            offset=o_train,
            family=model_version.family,
            link=self.link,
            preprocessor=model_version.preprocessor,
            alpha=model_version.alpha,
            l1_ratio=model_version.l1_ratio,
            use_cv=False,
            drop_reference=self.drop_reference,
        )
        pred_test = fold_mv.predict(X_test, offset=o_test)

        if dl_refit_base:
            bv = self._get_version(dl_base_version)
            base_mv = fit_model(
                X=X_train,
                y=y_train,
                variables=bv.variables,
                version_name="_cv_base_fold",
                weights=w_train,
                offset=o_train,
                family=bv.family,
                link=self.link,
                preprocessor=bv.preprocessor,
                alpha=bv.alpha,
                l1_ratio=bv.l1_ratio,
                use_cv=False,
                drop_reference=self.drop_reference,
            )
            base_preds_test = base_mv.predict(X_test, offset=o_test)
        else:
            base_preds_test = dl_base_preds[test_mask]

        _, fold_dl_score = self._get_dl_score(
            y_test,
            base_preds_test,
            pred_test,
            weights=w_test,
            n_buckets=10,
            deviation=dl_deviation,
        )

        return fold_dl_score, w_test.sum()

    def _monitor_baseline(
        self,
        version_names: list[str],
        base_version: str | None,
    ) -> tuple[list[str], np.ndarray | None, bool, str]:
        """Resolve double-lift baseline state for overfitting monitoring."""
        assert base_version is not None or self.base_version is not None, (
            "dl_base_version must be provided when metric_fn is "
            "'double_lift_score' and no base version is set."
        )
        resolved_base = base_version or self.base_version
        if resolved_base in self.model_versions:
            names = [name for name in version_names if name != resolved_base]
            predictions = self._prediction_source(resolved_base)
            return names, predictions, True, resolved_base
        if resolved_base in self.data.columns:
            predictions = self._prediction_source(resolved_base)
            return version_names, predictions, False, resolved_base
        self._get_version(resolved_base)
        return version_names, np.array([]), False, resolved_base

    def _monitor_train_metric(
        self,
        metric_name: str | None,
        model: ModelVersion,
        baseline_predictions: np.ndarray | None,
        deviation: str,
    ) -> float:
        """Calculate the in-sample metric for one monitor step."""
        if metric_name == "gini":
            return gini_coefficient(
                self._y_array,
                model.train_predictions,
                self._weights_array,
            )
        _, score = self._get_dl_score(
            self._y_array,
            baseline_predictions,
            model.train_predictions,
            weights=self._weights_array,
            n_buckets=10,
            deviation=deviation,
        )
        return score

    def _monitor_cv_metric(
        self,
        metric_name: str | None,
        model: ModelVersion,
        baseline_predictions: np.ndarray | None,
        refit_baseline: bool,
        base_version: str | None,
        deviation: str,
    ) -> float:
        """Calculate the exposure-weighted CV metric for one monitor step."""
        if self.cv_column is None:
            return self._monitor_train_metric(
                metric_name,
                model,
                baseline_predictions,
                deviation,
            )
        folds = self.data[self.cv_column].to_numpy()
        metrics: list[float] = []
        weights: list[float] = []
        for fold in np.unique(folds):
            train_mask = pl.Series(folds != fold)
            test_mask = folds == fold
            if metric_name == "gini":
                fold_metric, fold_weight = self._get_cv_metrics_gini(
                    train_mask,
                    test_mask,
                    model,
                )
            else:
                fold_metric, fold_weight = self._get_cv_metrics_dls(
                    train_mask,
                    test_mask,
                    baseline_predictions,
                    refit_baseline,
                    base_version,
                    model,
                    deviation,
                )
            metrics.append(fold_metric)
            weights.append(float(fold_weight))
        return float(np.average(metrics, weights=weights))

    def overfitting_monitor(
        self,
        version_names: list[str],
        metric_fn: Any | None = "double_lift_score",
        dl_base_version: str | None = None,
        dl_deviation: str = "absolute",
        show: bool = True,
    ) -> pl.DataFrame:
        """
        Track train vs CV metric across existing model versions.

        Requires ``cv_column`` to have been set at construction for CV metrics
        to differ from train metrics.

        Parameters
        ----------
        version_names : list of str
            Ordered list of registered version names to evaluate.
        metric_fn : {'gini', 'double_lift_score'}, optional
            Performance metric to track.
        dl_base_version : str, optional
            Baseline version or column name used when
            ``metric_fn='double_lift_score'``.  Falls back to
            ``self.base_version`` when ``None``.
        dl_deviation : {'absolute', 'relative'}
            Deviation metric for the double-lift score (default
            ``'absolute'``).
        show : bool
            Display the overfitting monitor chart (default ``True``)."""
        from .plots import overfitting_plot

        assert metric_fn in ("gini", "double_lift_score", None), (
            "metric_fn must be 'gini' or 'double_lift_score'"
        )
        dl_base_preds = None
        dl_refit_base = False
        if metric_fn == "double_lift_score":
            version_names, dl_base_preds, dl_refit_base, dl_base_version = (
                self._monitor_baseline(version_names, dl_base_version)
            )
        rows = []
        cumulative_vars: list[str] = []
        for index, version_name in enumerate(version_names):
            model = self._get_version(version_name)
            train_metric = self._monitor_train_metric(
                metric_fn,
                model,
                dl_base_preds,
                dl_deviation,
            )
            cv_metric = (
                train_metric
                if self.cv_column is None
                else self._monitor_cv_metric(
                    metric_fn,
                    model,
                    dl_base_preds,
                    dl_refit_base,
                    dl_base_version,
                    dl_deviation,
                )
            )
            new_vars = [v for v in model.variables if v not in cumulative_vars]
            cumulative_vars.extend(new_vars)
            rows.append(
                {
                    "step": index + 1,
                    "n_variables": len(model.variables),
                    "variables_added": version_name,
                    "train_metric": train_metric,
                    "cv_metric": cv_metric,
                    "gap": train_metric - cv_metric,
                }
            )

        monitor_df = pl.DataFrame(rows)
        if show:
            _fig = overfitting_plot(monitor_df)
            plt.show()
        return monitor_df

    # ── Statistical ──────────────────────────────────────────────────────────

    def vif_table(self, version: str | None = None) -> pl.DataFrame:
        """
        Compute VIF for each feature in a fitted model's design matrix.

        Parameters
        ----------
        version : str, optional
            Model version to analyse.  Uses current version if ``None``.
        """
        from .metrics import vif_table as _vif_table

        mv = self._get_version(version or self.current_version)
        # Reconstruct design matrix
        preprocessor = mv.preprocessor
        Xt = preprocessor.transform(self.data)
        feature_cols = [c for c in Xt.columns if c in mv.feature_names]
        design = Xt.select(feature_cols)
        return _vif_table(design)

    def bootstrap_metrics(
        self,
        version: str | None = None,
        metric_fns: Any | None = None,
        n_bootstrap: int = 500,
        ci: float = 0.95,
        show: bool = True,
    ) -> pl.DataFrame:
        """
        Bootstrap confidence intervals on model performance metrics.

        Parameters
        ----------
        version : str, optional
            Model version to evaluate.  Uses current version if ``None``.
        metric_fns : callable or list of callable, optional
            Metric functions ``(y_true, y_pred, weights) -> float``.  Uses
            default metrics when ``None``.
        n_bootstrap : int
            Number of bootstrap resamples (default 500).
        ci : float
            Confidence level for the interval (default 0.95).
        show : bool
            Display the confidence interval chart (default ``True``).
        """
        from .metrics import bootstrap_metrics as _bootstrap_metrics
        from .plots import bootstrap_ci_plot

        mv = self._get_version(version or self.current_version)
        result = _bootstrap_metrics(
            self._y_array,
            mv.train_predictions,
            weights=self._weights_array,
            metric_fns=metric_fns,
            n_bootstrap=n_bootstrap,
            ci=ci,
        )
        if show:
            _fig = bootstrap_ci_plot(
                result, title=f"Bootstrap CIs — {version or self.current_version}"
            )
            plt.show()
        return result

    def _fit_bootstrap_coefficients(
        self,
        model: ModelVersion,
        bootstrap_data: pl.DataFrame,
    ) -> dict[str, float]:
        """Refit one bootstrap sample and return its feature coefficients."""
        bootstrap_model = fit_model(
            X=bootstrap_data,
            y=bootstrap_data[self.target_col],
            variables=model.variables,
            version_name="_bootstrap",
            configs=self.variable_configs,
            preprocessor=model.preprocessor,
            weights=(
                bootstrap_data[self.weight_col].to_numpy().astype(float)
                if self.weight_col
                else None
            ),
            offset=(
                bootstrap_data[self.offset_col].to_numpy().astype(float)
                if self.offset_col
                else None
            ),
            family=model.family,
            link=self.link,
            tweedie_power=model.tweedie_power,
            alpha=model.alpha,
            l1_ratio=model.l1_ratio,
            use_cv=False,
            drop_reference=self.drop_reference,
        )
        return {
            row["feature"]: row["coefficient"]
            for row in bootstrap_model.coefficients.iter_rows(named=True)
        }

    @staticmethod
    def _append_bootstrap_coefficients(
        samples: dict[tuple[str, str], list[float]],
        baseline: dict[tuple[str, str], float],
        coefficients: dict[str, float] | None,
    ) -> None:
        """Append one fitted or fallback value for every relativity level."""
        for key, baseline_coefficient in baseline.items():
            variable, level = key
            feature = f"{variable}_{level}" if level != "(base)" else None
            value = (
                coefficients[feature]
                if coefficients is not None
                and feature is not None
                and feature in coefficients
                else baseline_coefficient
            )
            samples[key].append(value)

    @staticmethod
    def _bootstrap_relativity_table(
        baseline: dict[tuple[str, str], float],
        samples: dict[tuple[str, str], list[float]],
        alpha: float,
    ) -> pl.DataFrame:
        """Summarize bootstrapped coefficients on the relativity scale."""
        rows = []
        for (variable, level), coefficient in baseline.items():
            bootstrap_relativities = np.exp(np.asarray(samples[(variable, level)]))
            rows.append(
                {
                    "variable": variable,
                    "level": level,
                    "relativity": float(np.exp(coefficient)),
                    "ci_lower": float(np.quantile(bootstrap_relativities, alpha)),
                    "ci_upper": float(np.quantile(bootstrap_relativities, 1 - alpha)),
                    "std_error": float(np.std(bootstrap_relativities)),
                }
            )
        return pl.DataFrame(rows)

    def bootstrap_relativities(
        self,
        version: str | None = None,
        n_bootstrap: int = 200,
        ci: float = 0.95,
        random_state: int = 42,
        show: bool = False,
    ) -> pl.DataFrame:
        """
        Bootstrap CIs on each factor relativity by resampling and refitting.

        Parameters
        ----------
        version : str, optional
            Model version to bootstrap.  Uses current version if ``None``.
        n_bootstrap : int
            Number of bootstrap resamples (default 200).
        ci : float
            Confidence level for the interval (default 0.95).
        random_state : int
            Random seed for reproducibility (default 42).
        show : bool
            Display per-variable relativity CI plots (default ``False``).

        Returns
        -------
        pl.DataFrame
            Columns: ``variable``, ``level``, ``relativity``,
            ``ci_lower``, ``ci_upper``, ``std_error``.
        """
        mv = self._get_version(version or self.current_version)
        n = len(self.data)
        rng = np.random.RandomState(random_state)
        alpha = (1 - ci) / 2
        base_relativities = self.summary_table(version or self.current_version)
        base_coefs = {
            (row["variable"], row["level"]): row["train_coef"]
            for row in base_relativities.iter_rows(named=True)
        }
        boot_coefs: dict[tuple[str, str], list[float]] = {key: [] for key in base_coefs}

        for _ in range(n_bootstrap):
            bootstrap_data = self.data[rng.choice(n, n, replace=True)]
            try:
                coefficients = self._fit_bootstrap_coefficients(mv, bootstrap_data)
                self._append_bootstrap_coefficients(
                    boot_coefs,
                    base_coefs,
                    coefficients,
                )
            except Exception:
                self._append_bootstrap_coefficients(boot_coefs, base_coefs, None)

        result = self._bootstrap_relativity_table(base_coefs, boot_coefs, alpha)

        if show:
            from .plots import relativities_ci_plot

            for var in result["variable"].unique().to_list():
                _fig = relativities_ci_plot(result, var)
                plt.show()

        return result

    # ── Internal ──────────────────────────────────────────────────────────────

    @staticmethod
    def _strip_config(cfg: VariableConfig) -> VariableConfig:
        """Return cfg with binning/encoding removed so the Preprocessor outputs raw continuous values."""
        from dataclasses import replace as _dc_replace

        from .variable import MISSING_SENTINEL as _SENTINEL

        strip_kwargs: dict = {
            "n_bins": None,
            "bin_edges": None,
            "standardize": False,
            "degree": 1,
            "encoding": None,
        }
        if cfg.impute_strategy is None and (
            cfg.n_bins is not None or cfg.bin_edges is not None
        ):
            strip_kwargs.update(impute_strategy="constant", impute_value=_SENTINEL)
        return _dc_replace(cfg, **strip_kwargs)

    def _get_version(self, version: str) -> ModelVersion:
        if version not in self.model_versions:
            available = list(self.model_versions.keys())
            raise KeyError(f"Version '{version}' not found.  Available: {available}")
        return self.model_versions[version]
